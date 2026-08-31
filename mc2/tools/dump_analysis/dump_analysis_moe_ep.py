#  Copyright (c) 2025 Huawei Technologies Co., Ltd.
#  This program is free software, you can redistribute it and/or modify it under the terms and conditions of
#  CANN Open Software License Agreement Version 2.0 (the "License").
#  Please refer to the License for details. You may not use this file except in compliance with the License.
#  THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND, EITHER EXPRESS OR IMPLIED,
#  INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT, MERCHANTABILITY, OR FITNESS FOR A PARTICULAR PURPOSE.
#  See LICENSE in the root of the software repository for the full text of the License.
#

# !
#  \file dump_moe_ep_analysis.py
#  \brief
#

import logging
import os
import sys
from dataclasses import dataclass, field

import numpy as np
from openpyxl import Workbook

logging.basicConfig(
    level=logging.NOTSET,
    format="[%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()],
)


# Metadata 区固定大小 64KiB, 有效结构体 <=512B, 剩余保留
METADATA_SIZE = 64 * 1024
# Metadata 中 10 个 uint32 字段名, 按结构体偏移顺序排列
METADATA_FIELD_NAMES = [
    "layoutVersion",
    "nmt",
    "topK",
    "hidden",
    "epWorldSize",
    "localExpertNum",
    "aivNum",
    "networkMode",
    "serverNum",
    "rankNumPerServer",
]
# Metadata 中 7 个固定 Region 的名称
REGION_COUNT = 7
REGION_NAMES = [
    "PER_CORE_DIAG",
    "COUNT_NOTIFY",
    "EXPERT_COUNT",
    "DISPATCH_SLOT_STATE",
    "COMBINE_TOKEN_STATE",
    "COMBINE_STATE_TAIL",
    "HYBRID_SCALEOUT_STATUS",
]
# Region 名称到 dump 文件后缀的映射, 每个 Region 单独输出一个 .bin 文件
REGION_FILE_SUFFIX = {
    "METADATA": "metadata.bin",
    "PER_CORE_DIAG": "per_core_diag.bin",
    "COUNT_NOTIFY": "count_notify.bin",
    "EXPERT_COUNT": "expert_count.bin",
    "DISPATCH_SLOT_STATE": "dispatch_slot_state.bin",
    "COMBINE_TOKEN_STATE": "combine_token_state.bin",
    "COMBINE_STATE_TAIL": "combine_state_tail.bin",
    "HYBRID_SCALEOUT_STATUS": "hybrid_scaleout_status.bin",
}
# Per-core 诊断区: 每个 512B slot 包含 4 个 MoeEpCoreDiagRecord, 每个 record 64B
# MoeEpCoreDiagRecord 布局: uint64 opCnt(8B) + uint32 runPosition(4B) + uint32 epRankId(4B) + uint32 aivId(4B) + 保留(44B)
PER_CORE_SLOT_SIZE = 512
PER_CORE_RECORD_SIZE = 64
PER_CORE_RECORDS_PER_SLOT = 4
PER_CORE_OP_NAMES = ["dispatch", "dispatch_epilogue", "combine", "combine_epilogue"]


# 单卡分析结果, 封装 analyze_single_card 的返回值
@dataclass
class CardAnalysisResult:
    bs: int = 0
    core_count: int = 0
    context_data: list = field(default_factory=list)
    topk_idx_data: list = field(default_factory=list)
    per_core_data: list = field(default_factory=list)
    count_notify_data: list = field(default_factory=list)
    expert_count_data: list = field(default_factory=list)
    combine_token_state_data: list = field(default_factory=list)


# 全卡分析结果, 封装 save_to_excel 的参数
@dataclass
class AllAnalysisResult:
    all_metadata: dict = field(default_factory=dict)
    all_context: list = field(default_factory=list)
    all_topk_idx: list = field(default_factory=list)
    all_per_core: list = field(default_factory=list)
    all_count_notify: list = field(default_factory=list)
    all_expert_count: list = field(default_factory=list)
    all_combine_token_state: list = field(default_factory=list)
    card_count: int = 0


# 在卡目录中按后缀查找 Region 对应的 .bin 文件, 读取全部数据返回 uint8 数组
def read_region_file(target_path: str, suffix: str, log_filename: bool = False):
    for filename in os.listdir(target_path):
        if filename.endswith(suffix):
            file_path = os.path.join(target_path, filename)
            if log_filename:
                logging.info("读取 %s 文件数据", filename)
            with open(file_path, "rb") as f:
                return np.frombuffer(f.read(), dtype=np.uint8)
    logging.warning("未找到 %s 文件", suffix)
    return np.array([], dtype=np.uint8)


# 从 .metadata.bin 文件解析 Metadata
# 有效结构体布局: 10 个 uint32 字段 + 7 个 MoeEpDumpRegion(offset: uint64, size: uint64), 每个 region 占 4 个 uint32
def parse_metadata(target_path: str):
    metadata_dict = {name: 0 for name in METADATA_FIELD_NAMES}
    for name in REGION_NAMES:
        metadata_dict[name] = [0, 0]
    raw = read_region_file(target_path, REGION_FILE_SUFFIX["METADATA"])
    if raw.shape[0] < 512:
        logging.warning("metadata.bin 文件大小不足 512 字节, 无法解析 Metadata")
        return metadata_dict, False
    metadata_raw = raw[:512].view(np.uint32)
    # 解析 10 个 uint32 字段
    for i, name in enumerate(METADATA_FIELD_NAMES):
        metadata_dict[name] = int(metadata_raw[i])
    # 解析 7 个 Region 的 offset 和 size (每个 region: offset uint64 + size uint64 = 4 个 uint32)
    for i, name in enumerate(REGION_NAMES):
        base = 10 + i * 4
        offset = int(metadata_raw[base]) | (int(metadata_raw[base + 1]) << 32)
        size = int(metadata_raw[base + 2]) | (int(metadata_raw[base + 3]) << 32)
        metadata_dict[name] = [offset, size]
    return metadata_dict, True


# 遍历所有卡目录, 收集每张卡的 Metadata, 返回全局字典(每字段值为按卡序排列的数组)
def collect_all_metadata(card_dirs: list):
    all_metadata = {name: [] for name in METADATA_FIELD_NAMES + REGION_NAMES}
    for card_num, card_path in enumerate(card_dirs):
        if not os.path.isdir(card_path):
            logging.error("卡%d 路径不存在: %s, 跳过该卡", card_num, card_path)
            continue
        metadata_dict, ok = parse_metadata(card_path)
        if not ok:
            logging.error("卡%d Metadata 解析失败, 使用默认值填充", card_num)
            for name in METADATA_FIELD_NAMES:
                all_metadata[name].append(0)
            for name in REGION_NAMES:
                all_metadata[name].append([0, 0])
            continue
        for name in METADATA_FIELD_NAMES + REGION_NAMES:
            all_metadata[name].append(metadata_dict[name])
    return all_metadata


# 打印单卡 Metadata
def print_card_metadata(card_num: int, card_metadata: dict):
    logging.info("===== 卡%d Metadata 数据 =====", card_num)
    for name in METADATA_FIELD_NAMES + REGION_NAMES:
        logging.info("卡%d %s: %s", card_num, name, card_metadata.get(name))
    logging.info("===== 卡%d Metadata 打印结束 =====", card_num)


# 从 exception_info 解析后的 input.0.bin 中读取 context 数据
# context 对应 MoeCommContext 结构体:
#   uint32_t epRankId;                                  // 卡id
#   uint32_t rankSizePerServer;                         // epWorldSize
#   uint64_t epHcclBuffer[1024];                        // Win区地址
#   uint64_t hcommHandle[1024];                         // Handle地址
#   uint32_t channelsPerRank;                           // 每卡channel数
def get_context(target_path: str):
    for filename in os.listdir(target_path):
        if "input.0" in filename and filename.endswith("bin"):
            file_path = os.path.join(target_path, filename)
            context = np.fromfile(file_path, dtype=np.int32)
            logging.info("===== context 数据 =====")
            logging.info("context dtype: %s, size: %d", context.dtype, context.shape[0])
            raw = context.tobytes()
            # 前 2 个 uint32: epRankId, rankSizePerServer
            ep_rank_id = int(context[0])
            rank_size = min(int(context[1]) & 0xFFFFFFFF, 1024)
            logging.info("epRankId=%d, rankSizePerServer=%d", ep_rank_id, rank_size)
            # epHcclBuffer: 1024 个 uint64, 打印 rankSize 个 (Win区地址)
            offset = 8
            end = offset + rank_size * 8
            win_addrs = np.frombuffer(raw[offset:end], dtype=np.uint64)
            logging.info("epHcclBuffer(Win区地址)前%d个: %s", rank_size, win_addrs)
            # hcommHandle: 1024 个 uint64, 打印 rankSize 个 (Handle地址)
            offset = 8 + 1024 * 8
            end = offset + rank_size * 8
            handle_addrs = np.frombuffer(raw[offset:end], dtype=np.uint64)
            logging.info("hcommHandle(Handle地址)前%d个: %s", rank_size, handle_addrs)
            # channelsPerRank: 前面共 2 + 1024*2*2 = 4098 个 uint32
            chan_offset = (2 + 1024 * 4) * 4
            chan_end = chan_offset + 4
            channels = int.from_bytes(raw[chan_offset:chan_end], "little")
            logging.info("channelsPerRank=%d", channels)
            logging.info("===== context 打印结束 =====")
            return context
    logging.warning("未找到 input.0*bin 文件, 无法获取 context 数据")
    return None


# 从 input.2.bin 读取 topk_idx(int32), 按 Metadata 中的 topK reshape 为 (bs, k)
# 返回 reshape 后的数组和计算出的 bs
def get_topk_idx(target_path: str, topk: int, global_expert_num: int):
    for filename in os.listdir(target_path):
        if "input.2" in filename and filename.endswith("bin"):
            file_path = os.path.join(target_path, filename)
            topk_idx = np.fromfile(file_path, dtype=np.int32)
            bs = topk_idx.shape[0] // topk if topk > 0 else 0
            if topk > 0 and bs * topk != topk_idx.shape[0]:
                logging.warning(
                    "topk_idx 数据大小 %d 不能被 topK %d 整除, 截断为 %d",
                    topk_idx.shape[0],
                    topk,
                    bs,
                )
                topk_idx = topk_idx[: bs * topk]
            return topk_idx.reshape(bs, topk), bs
    logging.warning("未找到 input.2*bin 文件, 无法获取 topk_idx 数据")
    return None, 0


# 校验 topk_idx 值范围: 每个值应 >= 0 且 < global_expert_num, 异常项逐个打印
def check_topk_idx(topk_idx_reshape, global_expert_num: int):
    logging.info("===== topk_idx 数据 =====")
    logging.info("topk_idx 全部数据(bs, k):\n%s", topk_idx_reshape)
    mask = (topk_idx_reshape < 0) | (topk_idx_reshape >= global_expert_num)
    if np.any(mask):
        row_func, col_func = np.where(mask)
        logging.error(
            "检测到topk_idx中有%d个异常输入, 值应 >= 0 且 < 全局专家数%d",
            len(row_func),
            global_expert_num,
        )
        for r, c in zip(row_func, col_func):
            logging.error(
                "topk_idx下标[%d,%d]的值:%d超范围", r, c, topk_idx_reshape[r][c]
            )
    else:
        logging.info("未检测到topk_idx中有超范围的异常输入")
    logging.info("===== topk_idx 打印结束 =====")


# 分析 Per-core 诊断区: 从 .per_core_diag.bin 读取
# 布局: 100 个核 × 512B/核, 每核 512B slot 内含 4 个 MoeEpCoreDiagRecord(各 64B), 分别对应 4 个算子
# MoeEpCoreDiagRecord: uint64 opCnt + uint32 runPosition + uint32 epRankId + uint32 aivId
# 核数判断: 当某核 4 个 record 的 opCnt 和 runPosition 全为 0 时, 认为该核未使用
def analyze_per_core(target_path: str):
    raw_data = read_region_file(target_path, REGION_FILE_SUFFIX["PER_CORE_DIAG"])
    if raw_data.shape[0] == 0:
        logging.warning("Per-core 诊断区大小为0")
        return 0, []
    max_cores = raw_data.shape[0] // PER_CORE_SLOT_SIZE
    per_core_data = []
    core_count = max_cores
    for i in range(max_cores):
        slot_start = i * PER_CORE_SLOT_SIZE
        slot_end = slot_start + PER_CORE_SLOT_SIZE
        slot = raw_data[slot_start:slot_end]
        core_records = []
        all_zero = True
        for op_idx in range(PER_CORE_RECORDS_PER_SLOT):
            rec_start = op_idx * PER_CORE_RECORD_SIZE
            rec_end = rec_start + PER_CORE_RECORD_SIZE
            record = slot[rec_start:rec_end].view(np.uint32)
            op_count = int(record[0]) | (int(record[1]) << 32)
            run_pos = int(record[2])
            ep_rank_id = int(record[3])
            aiv_id = int(record[4])
            if op_count != 0 or run_pos != 0:
                all_zero = False
            core_records.append((op_count, run_pos, ep_rank_id, aiv_id))
        if all_zero:
            core_count = i
            break
        per_core_data.append(core_records)
    return core_count, per_core_data


# 打印 Per-core 分析结果: 按 op 分组, 每个 op 输出各核的 op_count/run_pos/epRankId/aivId 数组
def print_per_core(core_count: int, per_core_data: list):
    logging.info("使用核数: %d", core_count)
    for op_idx in range(PER_CORE_RECORDS_PER_SLOT):
        op_counts = [per_core_data[i][op_idx][0] for i in range(core_count)]
        run_poses = [per_core_data[i][op_idx][1] for i in range(core_count)]
        ep_rank_ids = [per_core_data[i][op_idx][2] for i in range(core_count)]
        aiv_ids = [per_core_data[i][op_idx][3] for i in range(core_count)]
        if all(c == 0 and r == 0 for c, r in zip(op_counts, run_poses)):
            continue
        logging.info("%s op_count: %s", PER_CORE_OP_NAMES[op_idx], op_counts)
        logging.info("%s run_pos: %s", PER_CORE_OP_NAMES[op_idx], run_poses)
        logging.info("%s epRankId: %s", PER_CORE_OP_NAMES[op_idx], ep_rank_ids)
        logging.info("%s aivId: %s", PER_CORE_OP_NAMES[op_idx], aiv_ids)


# 分析 Count Notify 区: 从 .count_notify.bin 读取
# 布局: epWorldSize 个 512B 块, 每块第一个 uint32 为该源 rank 发送的 token 计数
def analyze_count_notify(target_path: str, ep_world_size: int):
    raw_data = read_region_file(
        target_path, REGION_FILE_SUFFIX["COUNT_NOTIFY"], log_filename=True
    )
    if raw_data.shape[0] == 0:
        logging.warning("Count Notify 区大小为0")
        return []
    counts = []
    for i in range(ep_world_size):
        if (i + 1) * 512 > raw_data.shape[0]:
            logging.warning(
                "Count Notify 数据不足 %d 个 512B 块, 实际只有 %d 块", ep_world_size, i
            )
            break
        start = i * 512
        end = (i + 1) * 512
        block = raw_data[start:end].view(np.uint32)
        counts.append(int(block[0]))
    logging.info(
        "数据(epWorldSize=%d, 每 rank 一个 512B 块, 取第一个 uint32): %s",
        ep_world_size,
        counts,
    )
    return counts


# 分析 Expert Count 区: 从 .expert_count.bin 读取
# 布局: epWorldSize 个 512B 块, 每块包含 localExpertNum 个 uint32, 为该 rank 每个专家收到的 token 计数
def analyze_expert_count(target_path: str, ep_world_size: int, local_expert_num: int):
    raw_data = read_region_file(
        target_path, REGION_FILE_SUFFIX["EXPERT_COUNT"], log_filename=True
    )
    if raw_data.shape[0] == 0:
        logging.warning("Expert Count 区大小为0")
        return []
    expert_counts = []
    for i in range(ep_world_size):
        if (i + 1) * 512 > raw_data.shape[0]:
            logging.warning(
                "Expert Count 数据不足 %d 个 512B 块, 实际只有 %d 块", ep_world_size, i
            )
            break
        start = i * 512
        end = (i + 1) * 512
        block = raw_data[start:end].view(np.uint32)
        per_rank = [int(block[j]) for j in range(min(local_expert_num, 128))]
        expert_counts.append(per_rank)
    logging.info(
        "数据(epWorldSize=%d, 每 rank 一个 512B 块, 每 block 取 localExpertNum=%d 个 uint32): %s",
        ep_world_size,
        local_expert_num,
        expert_counts,
    )
    return expert_counts


# 分析 Combine Token State 区: 从 .combine_token_state.bin 读取
# 布局: N*K 个 512B 块(N=nmt, K=topK), 每块第一个 uint32 为该 (token, topk) 的回传完成状态
def analyze_combine_token_state(target_path: str, nmt: int, topk: int):
    raw_data = read_region_file(
        target_path, REGION_FILE_SUFFIX["COMBINE_TOKEN_STATE"], log_filename=True
    )
    if raw_data.shape[0] == 0:
        logging.warning("Combine Token State 区大小为0")
        return []
    total = nmt * topk
    states = []
    for i in range(total):
        if (i + 1) * 512 > raw_data.shape[0]:
            logging.warning(
                "Combine Token State 数据不足 %d 个 512B 块, 实际只有 %d 块", total, i
            )
            break
        start = i * 512
        end = (i + 1) * 512
        block = raw_data[start:end].view(np.uint32)
        states.append(int(block[0]))
    logging.info(
        "数据(nmt=%d * topK=%d = %d 个 512B 块, 取每块第一个 uint32): %s",
        nmt,
        topk,
        total,
        states,
    )
    return states


# 单卡分析主函数: 依次获取 context、topk_idx, 解析各 Region .bin 文件
# 返回 CardAnalysisResult 封装的单卡分析结果
def analyze_single_card(
    card_path: str, card_num: int, card_metadata: dict, global_expert_num: int
):
    logging.info("开始分析卡%d数据, dump数据路径: %s", card_num, card_path)
    print_card_metadata(card_num, card_metadata)
    logging.info("全局专家数: %d", global_expert_num)

    context = get_context(card_path)
    context_data = context.tolist() if context is not None else []

    topk = card_metadata.get("topK", 0) if card_metadata else 0
    topk_idx, bs = get_topk_idx(card_path, topk, global_expert_num)
    topk_idx_data = []
    if topk_idx is not None:
        logging.info("topk_idx bs: %d, k: %d", bs, topk)
        check_topk_idx(topk_idx, global_expert_num)
        topk_idx_data = topk_idx.tolist()

    ep_world_size = card_metadata.get("epWorldSize", 0) if card_metadata else 0
    local_expert_num = card_metadata.get("localExpertNum", 0) if card_metadata else 0
    nmt = card_metadata.get("nmt", 0) if card_metadata else 0

    core_count, per_core_data = analyze_per_core(card_path)
    print_per_core(core_count, per_core_data)
    count_notify_data = analyze_count_notify(card_path, ep_world_size)
    expert_count_data = analyze_expert_count(card_path, ep_world_size, local_expert_num)
    combine_token_state_data = analyze_combine_token_state(card_path, nmt, topk)

    return CardAnalysisResult(
        bs=bs,
        core_count=core_count,
        context_data=context_data,
        topk_idx_data=topk_idx_data,
        per_core_data=per_core_data,
        count_notify_data=count_notify_data,
        expert_count_data=expert_count_data,
        combine_token_state_data=combine_token_state_data,
    )


# 写 metadata sheet: 每卡一行, 10 个 uint32 字段 + 7 个 Region 拆为 offset/size 两列 + bs + core_count
def write_metadata_sheet(wb, all_metadata: dict, card_count: int):
    ws = wb.active
    ws.title = "metadata"
    region_offset_names = [r + "_offset" for r in REGION_NAMES]
    region_size_names = [r + "_size" for r in REGION_NAMES]
    all_field_names = (
        METADATA_FIELD_NAMES
        + region_offset_names
        + region_size_names
        + ["bs", "使用核数"]
    )
    ws.append(["card_num"] + all_field_names)
    for card_num in range(card_count):
        row = [card_num]
        for name in METADATA_FIELD_NAMES:
            val = (
                all_metadata[name][card_num]
                if card_num < len(all_metadata[name])
                else 0
            )
            row.append(val)
        for name in REGION_NAMES:
            val = (
                all_metadata[name][card_num]
                if card_num < len(all_metadata[name])
                else [0, 0]
            )
            offset, size = val if isinstance(val, list) else [val, 0]
            row.append(offset)
            row.append(size)
        row.append(
            all_metadata["bs"][card_num]
            if card_num < len(all_metadata.get("bs", []))
            else 0
        )
        row.append(
            all_metadata["core_count"][card_num]
            if card_num < len(all_metadata.get("core_count", []))
            else 0
        )
        ws.append(row)


# 写 context sheet: 按 MoeCommContext 结构体平铺, 每卡一行
# 列: card_num, epRankId, rankSizePerServer, epHcclBuffer_0..(rankSize-1), hcommHandle_0..(rankSize-1), channelsPerRank
def write_context_sheet(wb, all_context: list):
    ws = wb.create_sheet("context")
    max_rank_size = 0
    for context_data in all_context:
        if len(context_data) >= 2:
            rs = int(context_data[1]) & 0xFFFFFFFF
            if rs > max_rank_size:
                max_rank_size = rs
    header = ["card_num", "epRankId", "rankSizePerServer"]
    header += ["epHcclBuffer_" + str(i) for i in range(max_rank_size)]
    header += ["hcommHandle_" + str(i) for i in range(max_rank_size)]
    header += ["channelsPerRank"]
    ws.append(header)
    for card_num, context_data in enumerate(all_context):
        if len(context_data) < 4099:
            ws.append([card_num] + [None] * (3 + max_rank_size * 2 + 1))
            continue
        rank_size = int(context_data[1]) & 0xFFFFFFFF
        row = [card_num]
        row.append(int(context_data[0]) & 0xFFFFFFFF)
        row.append(rank_size)
        for i in range(max_rank_size):
            if i < rank_size:
                low = int(context_data[2 + i * 2]) & 0xFFFFFFFF
                high = int(context_data[2 + i * 2 + 1]) & 0xFFFFFFFF
                row.append(low | (high << 32))
            else:
                row.append(None)
        for i in range(max_rank_size):
            if i < rank_size:
                low = int(context_data[2050 + i * 2]) & 0xFFFFFFFF
                high = int(context_data[2050 + i * 2 + 1]) & 0xFFFFFFFF
                row.append(low | (high << 32))
            else:
                row.append(None)
        row.append(int(context_data[4098]) & 0xFFFFFFFF)
        ws.append(row)


# 写 topk_idx sheet: 每卡每(token, topk)一行, 包含 card_num, token_id, topk_id, expert_id
def write_topk_idx_sheet(wb, all_topk_idx: list):
    ws = wb.create_sheet("topk_idx")
    ws.append(["card_num", "token_id", "topk_id", "expert_id"])
    for card_num, topk_idx_data in enumerate(all_topk_idx):
        for token_id, row in enumerate(topk_idx_data):
            for topk_id, expert_id in enumerate(row):
                ws.append([card_num, token_id, topk_id, expert_id])


# 写 per_card_per_core sheet: 每核每 op 一行
# 列: card_num, core_count, core_id, op_index, op_name, op_count, run_pos, ep_rank_id, aiv_id
def write_per_core_sheet(wb, all_per_core: list):
    ws = wb.create_sheet("per_card_per_core")
    ws.append(
        [
            "card_num",
            "core_count",
            "core_id",
            "op_index",
            "op_name",
            "op_count",
            "run_pos",
            "ep_rank_id",
            "aiv_id",
        ]
    )
    for card_num, per_core_data in enumerate(all_per_core):
        core_count = len(per_core_data)
        for core_id, core_records in enumerate(per_core_data):
            for op_idx, (op_count, run_pos, ep_rank_id, aiv_id) in enumerate(
                core_records
            ):
                ws.append(
                    [
                        card_num,
                        core_count,
                        core_id,
                        op_idx,
                        PER_CORE_OP_NAMES[op_idx],
                        op_count,
                        run_pos,
                        ep_rank_id,
                        aiv_id,
                    ]
                )


# 写 count_notify sheet: 每卡每 rank 一行, 含 epWorldSize 便于核对行数
def write_count_notify_sheet(wb, all_count_notify: list, all_metadata: dict):
    ws = wb.create_sheet("count_notify")
    ws.append(["card_num", "epWorldSize", "rank_id", "count"])
    for card_num, count_notify_data in enumerate(all_count_notify):
        ep_ws = (
            all_metadata["epWorldSize"][card_num]
            if card_num < len(all_metadata["epWorldSize"])
            else 0
        )
        for rank_id, count in enumerate(count_notify_data):
            ws.append([card_num, ep_ws, rank_id, count])


# 写 expert_count sheet: 每卡每 rank 每 expert 一行, 含 localExpertNum 便于核对行数
def write_expert_count_sheet(wb, all_expert_count: list, all_metadata: dict):
    ws = wb.create_sheet("expert_count")
    ws.append(["card_num", "localExpertNum", "rank_id", "expert_id", "count"])
    for card_num, expert_count_data in enumerate(all_expert_count):
        local_expert_num = (
            all_metadata["localExpertNum"][card_num]
            if card_num < len(all_metadata["localExpertNum"])
            else 0
        )
        for rank_id, per_rank in enumerate(expert_count_data):
            for expert_id, count in enumerate(per_rank):
                ws.append([card_num, local_expert_num, rank_id, expert_id, count])


# 写 combine_token_state sheet: 每卡每 (token,topk) 一行, idx 为 0~N*K-1 的序号
def write_combine_token_state_sheet(wb, all_combine_token_state: list):
    ws = wb.create_sheet("combine_token_state")
    ws.append(["card_num", "idx", "token_state"])
    for card_num, combine_token_state_data in enumerate(all_combine_token_state):
        for idx, state in enumerate(combine_token_state_data):
            ws.append([card_num, idx, state])


# 将所有分析结果写入 Excel: 7 个 sheet(metadata/context/topk_idx/per_card_per_core/count_notify/expert_count/combine_token_state)
def save_to_excel(result: AllAnalysisResult):
    wb = Workbook()
    write_metadata_sheet(wb, result.all_metadata, result.card_count)
    write_context_sheet(wb, result.all_context)
    write_topk_idx_sheet(wb, result.all_topk_idx)
    write_per_core_sheet(wb, result.all_per_core)
    write_count_notify_sheet(wb, result.all_count_notify, result.all_metadata)
    write_expert_count_sheet(wb, result.all_expert_count, result.all_metadata)
    write_combine_token_state_sheet(wb, result.all_combine_token_state)
    wb.save("moe_ep_dump_analysis_result.xlsx")
    logging.info("全卡分析结果已归档至 moe_ep_dump_analysis_result.xlsx")


# ===== 主流程 =====
if __name__ == "__main__":
    # 1. 解析参数, 构建 card_dirs 列表
    if len(sys.argv) < 4:
        logging.error(
            "用法: python %s <dump_data_path> <file_num> <is_multi> [card_names]",
            sys.argv[0],
        )
        sys.exit(1)

    target_dir = os.path.realpath(sys.argv[1])
    file_num = int(sys.argv[2])
    is_multi = int(sys.argv[3])
    card_names = (
        [os.path.basename(name) for name in sys.argv[4].split(",")]
        if len(sys.argv) > 4
        else []
    )
    if not os.path.isdir(target_dir):
        logging.error("路径不存在: %s", target_dir)
        sys.exit(1)

    logging.info(
        "开始解析 dump 数据, 总卡数:%d, is_multi:%d, dump数据路径: %s",
        file_num,
        is_multi,
        target_dir,
    )

    if is_multi == 1:
        if card_names:
            card_dirs = [os.path.join(target_dir, name) for name in card_names]
        else:
            card_dirs = [os.path.join(target_dir, str(i)) for i in range(file_num)]
    else:
        card_dirs = [target_dir]

    # 2. 收集所有卡 Metadata, 计算全局专家数
    all_metadata = collect_all_metadata(card_dirs)
    logging.info("所有卡 Metadata 收集完成")

    global_expert_num = sum(all_metadata["localExpertNum"])
    logging.info("全局专家数: %d", global_expert_num)

    # 3. 逐卡执行单卡分析, 收集结果
    all_bs = []
    all_core_count = []
    all_context = []
    all_topk_idx = []
    all_per_core = []
    all_count_notify = []
    all_expert_count = []
    all_combine_token_state = []

    for card_num, card_path in enumerate(card_dirs):
        logging.info("--------------------------------------------")
        if not os.path.isdir(card_path):
            logging.error("卡%d 路径不存在: %s, 跳过该卡", card_num, card_path)
            all_bs.append(0)
            all_core_count.append(0)
            all_context.append([])
            all_topk_idx.append([])
            all_per_core.append([])
            all_count_notify.append([])
            all_expert_count.append([])
            all_combine_token_state.append([])
            continue
        card_metadata = {
            name: all_metadata[name][card_num]
            if card_num < len(all_metadata[name])
            else None
            for name in METADATA_FIELD_NAMES + REGION_NAMES
        }
        card_result = analyze_single_card(
            card_path, card_num, card_metadata, global_expert_num
        )
        all_bs.append(card_result.bs)
        all_core_count.append(card_result.core_count)
        all_context.append(card_result.context_data)
        all_topk_idx.append(card_result.topk_idx_data)
        all_per_core.append(card_result.per_core_data)
        all_count_notify.append(card_result.count_notify_data)
        all_expert_count.append(card_result.expert_count_data)
        all_combine_token_state.append(card_result.combine_token_state_data)

    all_metadata["bs"] = all_bs
    all_metadata["core_count"] = all_core_count

    # 4. 写入 Excel
    all_result = AllAnalysisResult(
        all_metadata=all_metadata,
        all_context=all_context,
        all_topk_idx=all_topk_idx,
        all_per_core=all_per_core,
        all_count_notify=all_count_notify,
        all_expert_count=all_expert_count,
        all_combine_token_state=all_combine_token_state,
        card_count=len(card_dirs),
    )
    save_to_excel(all_result)

    logging.info("--------------------------------------------")
    logging.info("所有卡数据解析完成, 多卡分析待实现")
