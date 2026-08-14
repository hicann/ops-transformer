#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# -----------------------------------------------------------------------------------------------------------
# Copyright (c) 2026 Huawei Technologies Co., Ltd.
# This program is free software, you can redistribute it and/or modify it under the terms and conditions of
# CANN Open Software License Agreement Version 2.0 (the "License").
# Please refer to the License for details. You may not use this file except in compliance with the License.
# THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND, EITHER EXPRESS OR IMPLIED,
# INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT, MERCHANTABILITY, OR FITNESS FOR A PARTICULAR PURPOSE.
# See LICENSE in the root of the software repository for the full text of the License.
# -----------------------------------------------------------------------------------------------------------

"""Regression tests for compare_batch_outputs.py."""

import csv
import json
import os
import subprocess
import sys
from pathlib import Path

import numpy as np
from openpyxl import load_workbook


SCRIPT = Path(__file__).with_name("compare_batch_outputs.py")
API_NAME = "qsmla_ttk_ops.quant_sparse_flash_mla_ttk"


class FixtureWriter:
    """Create small raw-bin E2E fixtures without requiring an NPU."""

    def __init__(self, root):
        self.root = root
        self.dump_dir = root / "dump"
        self.dump_dir.mkdir()
        self.case_csv = root / "cases.csv"
        self.result_csv = root / "result.csv"
        self.report = root / "report.json"
        self.excel = root / "report.xlsx"

    @staticmethod
    def batch_fields(slices, seed, sequence_slices=None):
        axes = (0,) if sequence_slices is None else (0, 1)
        slice_groups = (tuple(slices),)
        seed_groups = (tuple(seed for _ in slices),)
        if sequence_slices is not None:
            slice_groups += (tuple(sequence_slices),)
            seed_groups += (tuple(seed for _ in sequence_slices),)
        return {
            "batch_axis": repr((axes,)),
            "batch_slice_info": repr((slice_groups,)),
            "batch_seed": repr((seed_groups,)),
        }

    def case_row(self, testcase_name, batch_size, slices, seed):
        row = {
            "testcase_name": testcase_name,
            "is_enabled": "True",
            "api_name": API_NAME,
            "tensor_view_shapes": repr(((batch_size, 1, 1, 2),)),
            "tensor_dtypes": repr(("uint8",)),
            "input_data_ranges": repr(((-1, 1),)),
            "attributes": repr(
                {
                    "layout_q": "BSND",
                    "layout_kv": "BSND",
                    "q_datarange": [0.0, 0.001],
                    "ori_kv_datarange": [1.0, 2.0],
                    "cmp_kv_datarange": [-0.2, 0.2],
                    "seqused_q_values": [1] * batch_size,
                }
            ),
        }
        row.update(self.batch_fields(slices, seed))
        return row

    def tnd_case_row(self, testcase_name, batch_size, slices, seed):
        q_prefix = [2 * index for index in range(batch_size + 1)]
        ori_prefix = [3 * index for index in range(batch_size + 1)]
        cmp_prefix = list(range(batch_size + 1))
        row = {
            "testcase_name": testcase_name,
            "api_name": API_NAME,
            "tensor_view_shapes": repr(
                (
                    (q_prefix[-1], 1, 2),
                    (ori_prefix[-1], 1, 2),
                    (cmp_prefix[-1], 1, 2),
                    (1,),
                    (1,),
                    (1,),
                    None,
                    None,
                    None,
                    None,
                    (batch_size + 1,),
                    (batch_size + 1,),
                    (batch_size + 1,),
                    (batch_size,),
                    (batch_size,),
                    (batch_size,),
                    (batch_size,),
                )
            ),
            "tensor_dtypes": repr(
                (
                    "uint8",
                    "uint8",
                    "uint8",
                    "float32",
                    "float32",
                    "float32",
                    None,
                    None,
                    None,
                    None,
                    "int32",
                    "int32",
                    "int32",
                    "int32",
                    "int32",
                    "int32",
                    "int32",
                )
            ),
            "input_data_ranges": repr(
                (
                    (0, 255),
                    (0, 255),
                    (0, 255),
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    (0, q_prefix[-1]),
                    (0, ori_prefix[-1]),
                    (0, cmp_prefix[-1]),
                    (2, 2),
                    (3, 3),
                    (1, 1),
                    (0, 0),
                )
            ),
            "attributes": repr(
                {
                    "layout_q": "TND",
                    "layout_kv": "TND",
                    "block_num1": batch_size * 2,
                    "block_num2": batch_size,
                    "cu_seqlens_q_values": q_prefix,
                    "cu_seqlens_ori_kv_values": ori_prefix,
                    "cu_seqlens_cmp_kv_values": cmp_prefix,
                    "seqused_q_values": [2] * batch_size,
                    "seqused_ori_kv_values": [3] * batch_size,
                    "seqused_cmp_kv_values": [1] * batch_size,
                    "cmp_residual_kv_values": [0] * batch_size,
                }
            ),
        }
        row.update(self.batch_fields(slices, seed))
        return row

    def write_csv(self, path, rows):
        with path.open("w", encoding="utf-8", newline="") as stream:
            writer = csv.DictWriter(stream, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)

    @staticmethod
    def read_rows(path):
        with path.open("r", encoding="utf-8", newline="") as stream:
            return list(csv.DictReader(stream))

    def write_outputs(self, second_case_differs=False):
        relation = np.array([[[[1.0, 2.0]]]], dtype=np.float16)
        other = np.array([[[[5.0, 6.0]]]], dtype=np.float16)
        changed = np.array([[[[3.0, 4.0]]]], dtype=np.float16)
        first = np.concatenate((relation, relation), axis=0)
        second_relation = changed if second_case_differs else relation
        second = np.concatenate((relation, other, second_relation), axis=0)
        (self.dump_dir / "CASE_B2_output_0.bin").write_bytes(first.tobytes())
        (self.dump_dir / "CASE_B3_output_0.bin").write_bytes(second.tobytes())

    def write_fixture(self, second_case_differs=False):
        self.write_csv(
            self.case_csv,
            [
                self.case_row("CASE_B2", 2, ((0, 1, 1), (1, 2, 1)), 7001),
                self.case_row("CASE_B3", 3, ((0, 1, 1), (2, 3, 1)), 7001),
            ],
        )
        self.write_csv(
            self.result_csv,
            [
                {
                    "testcase_name": "CASE_B2",
                    "precision_status": "PASS",
                    "eager_precision": "100.0%,batch_intra=PASS",
                },
                {
                    "testcase_name": "CASE_B3",
                    "precision_status": "PASS",
                    "eager_precision": "100.0%,batch_intra=PASS",
                },
            ],
        )
        self.write_outputs(second_case_differs)

    def command(self, include_excel=True):
        command = [
            sys.executable,
            str(SCRIPT),
            "--csv",
            str(self.case_csv),
            "--result",
            str(self.result_csv),
            "--dump-dir",
            str(self.dump_dir),
            "--report",
            str(self.report),
        ]
        if include_excel:
            command.extend(("--excel", str(self.excel)))
        command.extend(("--require-intra-case", "--require-cross-case"))
        return command


def test_batch_output_comparator_accepts_same_and_cross_case(tmp_path):
    fixture = FixtureWriter(tmp_path)
    fixture.write_fixture()

    completed = subprocess.run(
        fixture.command(), check=False, capture_output=True, text=True
    )

    assert completed.returncode == 0, completed.stderr + completed.stdout
    report = json.loads(fixture.report.read_text(encoding="utf-8"))
    assert report["status"] == "PASS"
    assert report["group_count"] == 1
    assert report["groups"][0]["case_count"] == 2
    assert report["groups"][0]["sample_count"] == 4
    workbook = load_workbook(fixture.excel, data_only=True)
    assert workbook["Summary"]["B2"].value == "PASS"
    details = workbook["Relations"]
    assert details.max_row == 5
    assert "A2:A5" in {str(value) for value in details.merged_cells.ranges}
    assert "N2:N5" in {str(value) for value in details.merged_cells.ranges}
    assert details["F2"].value == "PASS"
    assert details["I2"].value
    assert report["groups"][0]["samples"][0]["dtype"] == "bfloat16"


def test_json_only_comparison_does_not_import_openpyxl(tmp_path):
    fixture = FixtureWriter(tmp_path)
    fixture.write_fixture()
    (tmp_path / "openpyxl.py").write_text(
        "raise RuntimeError('openpyxl must not load in JSON-only mode')\n",
        encoding="utf-8",
    )
    environment = os.environ.copy()
    environment["PYTHONPATH"] = os.pathsep.join(
        filter(
            None,
            (
                str(tmp_path),
                environment.get("PYTHONPATH"),
            ),
        )
    )

    completed = subprocess.run(
        fixture.command(include_excel=False),
        check=False,
        capture_output=True,
        text=True,
        env=environment,
    )

    assert completed.returncode == 0, completed.stderr + completed.stdout
    assert json.loads(fixture.report.read_text(encoding="utf-8"))["status"] == "PASS"
    assert not fixture.excel.exists()


def test_batch_output_comparator_rejects_cross_case_difference(tmp_path):
    fixture = FixtureWriter(tmp_path)
    fixture.write_fixture(second_case_differs=True)

    completed = subprocess.run(
        fixture.command(), check=False, capture_output=True, text=True
    )

    assert completed.returncode == 1
    report = json.loads(fixture.report.read_text(encoding="utf-8"))
    assert report["status"] == "FAIL"
    assert any("raw output differs" in error for error in report["groups"][0]["errors"])


def test_batch_output_comparator_rejects_truncated_zero_output(tmp_path):
    fixture = FixtureWriter(tmp_path)
    fixture.write_fixture()
    (fixture.dump_dir / "CASE_B2_output_0.bin").write_bytes(b"\0" * 6)
    (fixture.dump_dir / "CASE_B3_output_0.bin").write_bytes(b"\0" * 9)

    completed = subprocess.run(
        fixture.command(), check=False, capture_output=True, text=True
    )

    assert completed.returncode == 1
    assert "does not match shape/dtype expectation" in completed.stdout


def test_batch_output_comparator_rejects_missing_output(tmp_path):
    fixture = FixtureWriter(tmp_path)
    fixture.write_fixture()
    (fixture.dump_dir / "CASE_B3_output_0.bin").unlink()

    completed = subprocess.run(
        fixture.command(), check=False, capture_output=True, text=True
    )

    assert completed.returncode == 1
    assert "missing output dump" in completed.stdout


def test_batch_output_comparator_rejects_failed_ttk_case(tmp_path):
    fixture = FixtureWriter(tmp_path)
    fixture.write_fixture()
    fixture.write_csv(
        fixture.result_csv,
        [
            {
                "testcase_name": "CASE_B2",
                "precision_status": "PASS",
                "eager_precision": "100.0%,batch_intra=PASS",
            },
            {
                "testcase_name": "CASE_B3",
                "precision_status": "FAIL",
                "eager_precision": "100.0%,batch_intra=PASS",
            },
        ],
    )

    completed = subprocess.run(
        fixture.command(), check=False, capture_output=True, text=True
    )

    assert completed.returncode == 1
    assert "precision_status is 'FAIL', expected PASS" in completed.stdout


def test_batch_output_comparator_skips_disabled_case_without_result_or_dump(tmp_path):
    fixture = FixtureWriter(tmp_path)
    fixture.write_fixture()
    rows = fixture.read_rows(fixture.case_csv)
    disabled = fixture.case_row("CASE_DISABLED", 3, ((0, 1, 1), (2, 3, 1)), 7001)
    disabled["is_enabled"] = "False"
    fixture.write_csv(fixture.case_csv, [*rows, disabled])

    completed = subprocess.run(
        fixture.command(), check=False, capture_output=True, text=True
    )

    assert completed.returncode == 0, completed.stderr + completed.stdout
    report = json.loads(fixture.report.read_text(encoding="utf-8"))
    assert report["disabled_case_count"] == 1
    assert report["disabled_cases"] == ["CASE_DISABLED"]


def test_batch_output_comparator_keeps_relation_gate_after_disabling_member(tmp_path):
    fixture = FixtureWriter(tmp_path)
    fixture.write_fixture()
    case_rows = fixture.read_rows(fixture.case_csv)
    case_rows[1]["is_enabled"] = "False"
    fixture.write_csv(fixture.case_csv, case_rows)
    fixture.write_csv(fixture.result_csv, fixture.read_rows(fixture.result_csv)[:1])
    (fixture.dump_dir / "CASE_B3_output_0.bin").unlink()

    completed = subprocess.run(
        fixture.command(), check=False, capture_output=True, text=True
    )

    assert completed.returncode == 1
    assert "result CSV has no matching testcase" not in completed.stdout
    report = json.loads(fixture.report.read_text(encoding="utf-8"))
    assert report["disabled_cases"] == ["CASE_B3"]
    assert (
        "relation appears in fewer than two testcases" in report["groups"][0]["errors"]
    )


def test_batch_output_comparator_rejects_non_primary_output(tmp_path):
    fixture = FixtureWriter(tmp_path)
    fixture.write_fixture()

    completed = subprocess.run(
        [*fixture.command(), "--output-index", "1"],
        check=False,
        capture_output=True,
        text=True,
    )

    assert completed.returncode == 1
    assert "supports output index 0 only" in completed.stdout


def test_batch_output_comparator_requires_ttk_intra_case_result(tmp_path):
    fixture = FixtureWriter(tmp_path)
    fixture.write_fixture()
    rows = fixture.read_rows(fixture.result_csv)
    rows[1]["eager_precision"] = "100.0%"
    fixture.write_csv(fixture.result_csv, rows)

    completed = subprocess.run(
        fixture.command(), check=False, capture_output=True, text=True
    )

    assert completed.returncode == 1
    assert "eager_precision lacks batch_intra=PASS" in completed.stdout


def test_batch_output_comparator_rejects_different_input_context(tmp_path):
    fixture = FixtureWriter(tmp_path)
    fixture.write_fixture()
    rows = fixture.read_rows(fixture.case_csv)
    rows[1]["attributes"] = repr(
        {
            "layout_q": "BSND",
            "layout_kv": "BSND",
            "seqused_q_values": [2, 2, 2],
        }
    )
    fixture.write_csv(fixture.case_csv, rows)

    completed = subprocess.run(
        fixture.command(), check=False, capture_output=True, text=True
    )

    assert completed.returncode == 1
    report = json.loads(fixture.report.read_text(encoding="utf-8"))
    assert any(
        "input context differs" in error for error in report["groups"][0]["errors"]
    )


def test_batch_output_comparator_requires_cross_case_when_requested(tmp_path):
    fixture = FixtureWriter(tmp_path)
    fixture.write_fixture()
    fixture.write_csv(
        fixture.case_csv,
        [
            fixture.case_row("CASE_B2", 2, ((0, 1, 1), (1, 2, 1)), 7002),
            fixture.case_row("CASE_B3", 3, ((0, 1, 1), (2, 3, 1)), 7003),
        ],
    )

    completed = subprocess.run(
        fixture.command(), check=False, capture_output=True, text=True
    )

    assert completed.returncode == 1
    report = json.loads(fixture.report.read_text(encoding="utf-8"))
    assert all(
        "relation appears in fewer than two testcases" in group["errors"]
        for group in report["groups"]
    )


def test_batch_output_comparator_accepts_tnd_logical_relations(tmp_path):
    fixture = FixtureWriter(tmp_path)
    fixture.write_csv(
        fixture.case_csv,
        [
            fixture.tnd_case_row("TND_B3", 3, ((0, 1, 1), (2, 3, 1)), 7011),
            fixture.tnd_case_row("TND_B4", 4, ((0, 1, 1), (3, 4, 1)), 7011),
        ],
    )
    fixture.write_csv(
        fixture.result_csv,
        [
            {
                "testcase_name": name,
                "precision_status": "PASS",
                "eager_precision": "100.0%,batch_intra=PASS",
            }
            for name in ("TND_B3", "TND_B4")
        ],
    )
    relation = np.array([[[1.0, 2.0]], [[3.0, 4.0]]], dtype=np.float16)
    for name, extent, target in (
        ("TND_B3", 6, (4, 6)),
        ("TND_B4", 8, (6, 8)),
    ):
        output = np.zeros((extent, 1, 2), dtype=np.float16)
        output[0:2] = relation
        output[target[0] : target[1]] = relation
        (fixture.dump_dir / f"{name}_output_0.bin").write_bytes(output.tobytes())

    completed = subprocess.run(
        fixture.command(), check=False, capture_output=True, text=True
    )

    assert completed.returncode == 0, completed.stderr + completed.stdout
