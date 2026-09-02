#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# -----------------------------------------------------------------------------------------------------------
# Copyright (c) 2026 Huawei Technologies Co., Ltd.
# This program is free software, you can redistribute it and/or modify it under the terms and conditions of
# CANN Open Software License Agreement Version 2.0 (the "License").
# Please refer to the License for details. You may not use this file except in compliance with the License.
# THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND, EITHER EXPRESS OR IMPLIED,
# INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT, MERCHANTABILITY, OR FITNESS FOR A PARTICULAR PURPOSE.
# See LICENSE in the root of the software repository for the full text of the License.
# -----------------------------------------------------------------------------------------------------------

"""Compare E2E batch-consistency output bins after all TTK cases have finished."""

import argparse
import ast
import csv
import hashlib
import json
import os
import tempfile
from collections import Counter, defaultdict
from pathlib import Path


class BatchOutputComparator:
    """Validate same-case and cross-case relations from raw output bins."""

    PRIMARY_OUTPUT_INDEX = 0
    PRIMARY_OUTPUT_DTYPE = "bfloat16"
    PREFIX_LENGTH_INPUT_INDEXES = frozenset((10, 11, 12))
    PREFIX_ATTRIBUTE_NAMES = (
        "cu_seqlens_q_values",
        "cu_seqlens_ori_kv_values",
        "cu_seqlens_cmp_kv_values",
    )
    STATIC_SEQUENCE_ATTRIBUTE_NAMES = frozenset(
        (
            "q_datarange",
            "ori_kv_datarange",
            "cmp_kv_datarange",
        )
    )
    BLOCK_COUNT_ATTRIBUTE_NAMES = frozenset(("block_num1", "block_num2"))

    BYTE_WIDTHS = {
        "bool": 1,
        "uint8": 1,
        "int8": 1,
        "float8_e4m3fn": 1,
        "float8_e5m2": 1,
        "uint16": 2,
        "int16": 2,
        "float16": 2,
        "bfloat16": 2,
        "uint32": 4,
        "int32": 4,
        "float32": 4,
        "uint64": 8,
        "int64": 8,
        "float64": 8,
        "complex64": 8,
        "complex128": 16,
    }

    def __init__(
        self,
        case_csv,
        result_csv,
        dump_dir,
        report_path,
        output_index,
        require_intra_case,
        require_cross_case,
        excel_path=None,
    ):
        self.case_csv = Path(case_csv)
        self.result_csv = Path(result_csv)
        self.dump_dir = Path(dump_dir)
        self.report_path = Path(report_path)
        self.output_index = output_index
        self.require_intra_case = require_intra_case
        self.require_cross_case = require_cross_case
        self.excel_path = Path(excel_path) if excel_path else None

    @staticmethod
    def read_rows(path):
        with path.open("r", encoding="utf-8", newline="") as stream:
            return list(csv.DictReader(stream))

    @staticmethod
    def parse_cell(row, name, default=None):
        value = row.get(name)
        if value is None or value == "":
            return default
        try:
            return ast.literal_eval(value)
        except (SyntaxError, ValueError) as error:
            testcase_name = row.get("testcase_name", "<unknown>")
            raise ValueError(f"{testcase_name}: invalid {name}: {error}") from error

    @staticmethod
    def case_is_enabled(row):
        """Mirror TTK's CSV boolean semantics; an absent value defaults to enabled."""
        value = row.get("is_enabled")
        if value is None or not value.strip():
            return True
        normalized = value.strip()
        if normalized.upper() in ("TRUE", "FALSE"):
            normalized = normalized.title()
        try:
            return bool(ast.literal_eval(normalized))
        except (SyntaxError, ValueError) as error:
            testcase_name = row.get("testcase_name", "<unknown>")
            raise ValueError(f"{testcase_name}: invalid is_enabled: {error}") from error

    @staticmethod
    def normalize_value(value):
        if isinstance(value, dict):
            return {
                str(key): BatchOutputComparator.normalize_value(item)
                for key, item in sorted(value.items(), key=lambda item: str(item[0]))
            }
        if isinstance(value, (list, tuple)):
            return [BatchOutputComparator.normalize_value(item) for item in value]
        return value

    def normalize_shapes(self, shapes, dynamic_extents):
        normalized = []
        for index, shape in enumerate(shapes):
            if shape is None:
                normalized.append(None)
                continue
            value = tuple(shape)
            if index in self.PREFIX_LENGTH_INPUT_INDEXES:
                normalized.append(["PREFIX", *value[1:]])
            elif value and value[0] in dynamic_extents:
                normalized.append([dynamic_extents[value[0]], *value[1:]])
            else:
                normalized.append(list(value))
        return normalized

    def normalize_ranges(self, ranges):
        normalized = []
        for index, value in enumerate(ranges):
            # Prefix storage range scales with B + 1 and is not relation identity.
            normalized.append(
                None
                if index in self.PREFIX_LENGTH_INPUT_INDEXES
                else self.normalize_value(value)
            )
        return normalized

    @staticmethod
    def normalize_prefix(value, start, stop):
        selected = [int(item) for item in value[start : stop + 1]]
        base = selected[0]
        return [item - base for item in selected]

    def normalize_attributes(
        self, attributes, start, stop, batch_size, relation_token_lengths
    ):
        normalized = {}
        for key, value in attributes.items():
            if key in self.PREFIX_ATTRIBUTE_NAMES and isinstance(value, (list, tuple)):
                normalized[key] = self.normalize_prefix(value, start, stop)
            elif key in self.STATIC_SEQUENCE_ATTRIBUTE_NAMES:
                normalized[key] = self.normalize_value(value)
            elif (
                key in self.BLOCK_COUNT_ATTRIBUTE_NAMES
                and attributes.get("layout_kv") != "PA_BBND"
            ):
                continue
            elif isinstance(value, (list, tuple)) and len(value) == batch_size:
                normalized[key] = self.normalize_value(value[start:stop])
            elif key == "B":
                normalized[key] = stop - start
            elif key == "T1" and relation_token_lengths.get("q") is not None:
                normalized[key] = relation_token_lengths["q"]
            elif key == "T2" and relation_token_lengths.get("ori_kv") is not None:
                normalized[key] = relation_token_lengths["ori_kv"]
            elif key == "T3" and relation_token_lengths.get("cmp_kv") is not None:
                normalized[key] = relation_token_lengths["cmp_kv"]
            else:
                normalized[key] = self.normalize_value(value)
        return normalized

    def layout_context(self, row, output_extent):
        attributes = self.parse_cell(row, "attributes", {})
        layout_q = attributes.get("layout_q", "BSND")
        if layout_q not in ("BSND", "TND"):
            raise ValueError(
                f"{row['testcase_name']}: unsupported layout_q={layout_q!r}"
            )
        prefixes = {}
        for name in self.PREFIX_ATTRIBUTE_NAMES:
            value = attributes.get(name)
            if value is not None:
                prefixes[name] = [int(item) for item in value]
        if layout_q == "TND":
            q_prefix = prefixes.get("cu_seqlens_q_values")
            if q_prefix is None:
                raise ValueError(
                    f"{row['testcase_name']}: TND requires cu_seqlens_q_values"
                )
            if q_prefix[0] != 0 or q_prefix[-1] != output_extent:
                raise ValueError(
                    f"{row['testcase_name']}: cu_seqlens_q_values must span output T"
                )
            batch_size = len(q_prefix) - 1
        else:
            q_prefix = None
            batch_size = output_extent
        for name, value in prefixes.items():
            if len(value) != batch_size + 1 or value[0] != 0:
                raise ValueError(
                    f"{row['testcase_name']}: {name} must contain B + 1 prefix values"
                )
            if any(right <= left for left, right in zip(value, value[1:])):
                raise ValueError(
                    f"{row['testcase_name']}: {name} must be strictly increasing"
                )
        return attributes, batch_size, q_prefix, prefixes

    @staticmethod
    def logical_batch_slice(testcase_name, start, stop, q_prefix):
        if q_prefix is None:
            return start, stop
        boundaries = {offset: index for index, offset in enumerate(q_prefix)}
        if start not in boundaries or stop not in boundaries:
            raise ValueError(
                f"{testcase_name}: TND q slice {(start, stop, 1)!r} must align "
                "with complete cu_seqlens_q intervals"
            )
        return boundaries[start], boundaries[stop]

    @staticmethod
    def relation_token_lengths(prefixes, start, stop):
        result = {"q": None, "ori_kv": None, "cmp_kv": None}
        names = {
            "cu_seqlens_q_values": "q",
            "cu_seqlens_ori_kv_values": "ori_kv",
            "cu_seqlens_cmp_kv_values": "cmp_kv",
        }
        for name, target in names.items():
            value = prefixes.get(name)
            if value is not None:
                result[target] = value[stop] - value[start]
        return result

    @staticmethod
    def dynamic_extents(batch_size, prefixes):
        labels = {batch_size: {"B"}, batch_size + 1: {"B+1"}}
        for name, value in prefixes.items():
            label = name.removeprefix("cu_seqlens_").removesuffix("_values").upper()
            labels.setdefault(value[-1], set()).add(f"T_{label}")
        return {extent: "/".join(sorted(value)) for extent, value in labels.items()}

    @classmethod
    def expected_byte_count(cls, output_shape, output_dtype):
        dtype_name = str(output_dtype).lower()
        if dtype_name.startswith("torch."):
            dtype_name = dtype_name[len("torch.") :]
        byte_width = cls.BYTE_WIDTHS.get(dtype_name)
        if byte_width is None:
            raise ValueError(f"unsupported output dtype for raw bin: {output_dtype!r}")
        element_count = 1
        for dimension in output_shape:
            if not isinstance(dimension, int) or dimension <= 0:
                raise ValueError(f"invalid output shape for raw bin: {output_shape!r}")
            element_count *= dimension
        return element_count * byte_width

    @staticmethod
    def parse_slice(testcase_name, axis, value):
        if not isinstance(value, (tuple, list)) or len(value) != 3:
            raise ValueError(
                f"{testcase_name}: invalid logical axis {axis} slice {value!r}"
            )
        if not all(isinstance(item, int) for item in value):
            raise ValueError(f"{testcase_name}: logical slice must contain integers")
        start, stop, step = (int(item) for item in value)
        if step != 1 or start < 0 or start >= stop:
            raise ValueError(
                f"{testcase_name}: logical slice must be non-empty and contiguous"
            )
        return start, stop, step

    @staticmethod
    def validate_disjoint_relations(testcase_name, relations):
        """Reject relation samples that select the same logical q positions."""
        for index, relation in enumerate(relations):
            for candidate in relations[index + 1 :]:
                if relation["axes"] != candidate["axes"]:
                    continue
                batch_slice = relation["slices"][0]
                candidate_batch = candidate["slices"][0]
                if (
                    batch_slice[1] <= candidate_batch[0]
                    or candidate_batch[1] <= batch_slice[0]
                ):
                    continue
                if relation["axes"] == (0,):
                    raise ValueError(
                        f"{testcase_name}: relation samples overlap logical B positions"
                    )
                sequence_slice = relation["slices"][1]
                candidate_sequence = candidate["slices"][1]
                if not (
                    sequence_slice[1] <= candidate_sequence[0]
                    or candidate_sequence[1] <= sequence_slice[0]
                ):
                    raise ValueError(
                        f"{testcase_name}: relation samples overlap logical q positions"
                    )

    def parse_relations(self, row):
        batch_axis = self.parse_cell(row, "batch_axis")
        batch_slices = self.parse_cell(row, "batch_slice_info")
        batch_seed = self.parse_cell(row, "batch_seed")
        fields = (batch_axis, batch_slices, batch_seed)
        if any(field is None for field in fields):
            raise ValueError(
                f"{row['testcase_name']}: batch_axis, batch_slice_info and batch_seed are required"
            )
        if not (len(batch_axis) == len(batch_slices) == len(batch_seed)):
            raise ValueError(
                f"{row['testcase_name']}: batch metadata top-level counts differ"
            )
        if not batch_axis or tuple(batch_axis[0]) not in ((0,), (0, 1)):
            raise ValueError(
                f"{row['testcase_name']}: q must use logical axes (0,) or (0, 1)"
            )
        if batch_slices[0] is None or batch_seed[0] is None:
            raise ValueError(
                f"{row['testcase_name']}: q slices and q seeds are required"
            )
        if any(value is not None for value in batch_slices[1:]):
            raise ValueError(f"{row['testcase_name']}: only q relations are supported")
        if any(value is not None for value in batch_seed[1:]):
            raise ValueError(
                f"{row['testcase_name']}: only q relation seeds are supported"
            )
        axes = tuple(batch_axis[0])
        axis_slices = batch_slices[0]
        axis_seeds = batch_seed[0]
        if len(axis_slices) != len(axes) or len(axis_seeds) != len(axes):
            raise ValueError(
                f"{row['testcase_name']}: q groups must match logical axes"
            )
        sample_count = len(axis_slices[0])
        if not sample_count or any(
            len(values) != sample_count for values in (*axis_slices, *axis_seeds)
        ):
            raise ValueError(
                f"{row['testcase_name']}: q axis sample counts differ or are empty"
            )

        relations = []
        for sample_index in range(sample_count):
            slices = []
            seed = None
            for axis_group, axis in enumerate(axes):
                slices.append(
                    self.parse_slice(
                        row["testcase_name"],
                        axis,
                        axis_slices[axis_group][sample_index],
                    )
                )
                seed_value = axis_seeds[axis_group][sample_index]
                if not isinstance(seed_value, int):
                    raise ValueError(
                        f"{row['testcase_name']}: q seed must be an integer"
                    )
                if seed is not None and seed != seed_value:
                    raise ValueError(
                        f"{row['testcase_name']}: logical B and S must use the same seed"
                    )
                seed = int(seed_value)
            if axes == (0, 1) and slices[0][1] - slices[0][0] != 1:
                raise ValueError(
                    f"{row['testcase_name']}: logical (B,S) requires one B per sample"
                )
            relations.append({"axes": axes, "slices": tuple(slices), "seed": seed})

        self.validate_disjoint_relations(row["testcase_name"], relations)

        if axes == (0, 1) and len({item["slices"][1] for item in relations}) > 1:
            attributes = self.parse_cell(row, "attributes", {})
            masks = [attributes.get("ori_mask_mode")]
            if attributes.get("has_cmp_kv", True):
                masks.append(attributes.get("cmp_mask_mode"))
            if any(value not in (None, 0) for value in masks):
                raise ValueError(
                    f"{row['testcase_name']}: shifted logical S relations require no-mask mode"
                )
        return relations

    def resolve_output_metadata(self, shapes):
        """Return output-0 metadata from the QSMLA operator contract."""
        if self.output_index != self.PRIMARY_OUTPUT_INDEX:
            raise ValueError("QSMLA batch consistency supports output index 0 only")
        if not shapes or shapes[0] is None:
            raise ValueError("QSMLA batch consistency requires q tensor metadata")
        # QSMLA output-0 shape follows q, but its dtype is BF16 rather than q's uint8.
        return tuple(shapes[0]), self.PRIMARY_OUTPUT_DTYPE

    def build_context(self, row, relation, output_shape, output_dtype):
        shapes = self.parse_cell(row, "tensor_view_shapes")
        dtypes = self.parse_cell(row, "tensor_dtypes")
        ranges = self.parse_cell(row, "input_data_ranges", ())
        output_extent = output_shape[0]
        attributes, batch_size, q_prefix, prefixes = self.layout_context(
            row, output_extent
        )
        batch_slice = relation["slices"][0]
        batch_start, batch_stop, _ = batch_slice
        sequence_slice = relation["slices"][1] if relation["axes"] == (0, 1) else None
        token_lengths = self.relation_token_lengths(prefixes, batch_start, batch_stop)
        if sequence_slice is not None:
            token_lengths["q"] = sequence_slice[1] - sequence_slice[0]
        feature_shape = (
            output_shape[2:]
            if attributes.get("layout_q", "BSND") == "BSND"
            and sequence_slice is not None
            else output_shape[1:]
        )
        sequence_context = None
        if sequence_slice is not None:
            masks = [attributes.get("ori_mask_mode")]
            if attributes.get("has_cmp_kv", True):
                masks.append(attributes.get("cmp_mask_mode"))
            sequence_context = (
                [0, sequence_slice[1] - sequence_slice[0]]
                if all(value in (None, 0) for value in masks)
                else list(sequence_slice[:2])
            )
        return {
            "api_name": row.get("api_name"),
            "layout_q": attributes.get("layout_q", "BSND"),
            "layout_kv": attributes.get("layout_kv", "BSND"),
            "relation_batch_count": batch_stop - batch_start,
            "relation_axes": list(relation["axes"]),
            "relation_sequence_slice": sequence_context,
            "relation_token_lengths": token_lengths,
            "input_shapes_without_relation": self.normalize_shapes(
                shapes, self.dynamic_extents(batch_size, prefixes)
            ),
            "input_dtypes": self.normalize_value(dtypes),
            "input_ranges": self.normalize_ranges(ranges),
            "attributes": self.normalize_attributes(
                attributes, batch_start, batch_stop, batch_size, token_lengths
            ),
            "output_feature_shape": list(feature_shape),
            "output_dtype": str(output_dtype),
        }

    def extract_output(self, row, relation, output_shape, output_dtype, output_bytes):
        attributes, batch_size, q_prefix, _prefixes = self.layout_context(
            row, output_shape[0]
        )
        batch_slice = relation["slices"][0]
        batch_start, batch_stop, _ = batch_slice
        if batch_stop > batch_size:
            raise ValueError(
                f"{row['testcase_name']}: logical B slice exceeds B={batch_size}"
            )
        sequence_slice = relation["slices"][1] if relation["axes"] == (0, 1) else None
        byte_width = self.expected_byte_count((1,), output_dtype)
        if attributes.get("layout_q", "BSND") == "BSND":
            if len(output_shape) < 2:
                raise ValueError(f"{row['testcase_name']}: BSND output has no S axis")
            sequence_extent = output_shape[1]
            trailing_count = 1
            for dimension in output_shape[2:]:
                trailing_count *= dimension
            if sequence_slice is None:
                element_start = batch_start * sequence_extent * trailing_count
                element_stop = batch_stop * sequence_extent * trailing_count
                sample_shape = [batch_stop - batch_start, *output_shape[1:]]
            else:
                if sequence_slice[1] > sequence_extent:
                    raise ValueError(
                        f"{row['testcase_name']}: logical S slice exceeds output S"
                    )
                element_start = (
                    batch_start * sequence_extent + sequence_slice[0]
                ) * trailing_count
                element_stop = (
                    batch_start * sequence_extent + sequence_slice[1]
                ) * trailing_count
                sample_shape = [
                    1,
                    sequence_slice[1] - sequence_slice[0],
                    *output_shape[2:],
                ]
        else:
            trailing_count = 1
            for dimension in output_shape[1:]:
                trailing_count *= dimension
            if sequence_slice is None:
                token_start, token_stop = q_prefix[batch_start], q_prefix[batch_stop]
            else:
                token_start = q_prefix[batch_start] + sequence_slice[0]
                token_stop = q_prefix[batch_start] + sequence_slice[1]
                if token_stop > q_prefix[batch_start + 1]:
                    raise ValueError(
                        f"{row['testcase_name']}: logical S slice exceeds TND interval"
                    )
            element_start = token_start * trailing_count
            element_stop = token_stop * trailing_count
            sample_shape = [token_stop - token_start, *output_shape[1:]]
        return (
            output_bytes[element_start * byte_width : element_stop * byte_width],
            sample_shape,
        )

    def build_samples(self, case_rows, result_rows):
        result_by_case = {}
        for row in result_rows:
            testcase_name = row.get("testcase_name")
            if not testcase_name:
                continue
            if testcase_name in result_by_case:
                raise ValueError(f"duplicate result row for testcase {testcase_name}")
            result_by_case[testcase_name] = row

        samples = []
        for row in case_rows:
            testcase_name = row.get("testcase_name")
            if not testcase_name:
                raise ValueError("case CSV has an empty testcase_name")
            if row.get("batch_axis") in (None, ""):
                continue
            result = result_by_case.get(testcase_name)
            if result is None:
                raise ValueError(
                    f"{testcase_name}: result CSV has no matching testcase"
                )
            if result.get("precision_status") != "PASS":
                raise ValueError(
                    f"{testcase_name}: precision_status is "
                    f"{result.get('precision_status')!r}, expected PASS"
                )
            eager_precision = result.get("eager_precision") or ""
            if "NO_OUTPU" in eager_precision:
                raise ValueError(
                    f"{testcase_name}: eager_precision reports no output; "
                    "raw-byte batch comparison is invalid"
                )
            if self.require_intra_case and "batch_intra=PASS" not in eager_precision:
                raise ValueError(
                    f"{testcase_name}: eager_precision lacks batch_intra=PASS"
                )

            shapes = self.parse_cell(row, "tensor_view_shapes")
            dtypes = self.parse_cell(row, "tensor_dtypes")
            if not dtypes:
                raise ValueError(f"{testcase_name}: tensor dtype metadata is required")
            output_shape, output_dtype = self.resolve_output_metadata(shapes)
            if not output_shape or output_shape[0] <= 0:
                raise ValueError(
                    f"{testcase_name}: output 0 requires a non-empty batch axis"
                )
            output_path = (
                self.dump_dir / f"{testcase_name}_output_{self.output_index}.bin"
            )
            if not output_path.is_file():
                raise ValueError(f"{testcase_name}: missing output dump {output_path}")
            output_bytes = output_path.read_bytes()
            expected_bytes = self.expected_byte_count(output_shape, output_dtype)
            if len(output_bytes) != expected_bytes:
                raise ValueError(
                    f"{testcase_name}: output byte count {len(output_bytes)} does not match "
                    f"shape/dtype expectation {expected_bytes}"
                )
            for logical_relation in self.parse_relations(row):
                value, sample_shape = self.extract_output(
                    row, logical_relation, output_shape, output_dtype, output_bytes
                )
                logical_sizes = tuple(
                    stop - start for start, stop, _step in logical_relation["slices"]
                )
                relation = (
                    self.output_index,
                    logical_relation["axes"],
                    logical_relation["seed"],
                    logical_sizes,
                )
                samples.append(
                    {
                        "testcase_name": testcase_name,
                        "relation": relation,
                        "slice": {
                            "B": list(logical_relation["slices"][0]),
                            "S": (
                                list(logical_relation["slices"][1])
                                if logical_relation["axes"] == (0, 1)
                                else None
                            ),
                        },
                        "shape": sample_shape,
                        "dtype": str(output_dtype),
                        "context": self.build_context(
                            row, logical_relation, output_shape, output_dtype
                        ),
                        "digest": hashlib.sha256(value).hexdigest(),
                        "value": value,
                    }
                )
        if not samples:
            raise ValueError("case CSV has no batch-consistency relations")
        return samples

    def compare_group(self, relation, samples):
        reference = samples[0]
        errors = []
        for sample in samples[1:]:
            if (
                sample["shape"] != reference["shape"]
                or sample["dtype"] != reference["dtype"]
            ):
                errors.append(
                    f"{sample['testcase_name']}: relation shape/dtype differs from "
                    f"{reference['testcase_name']}"
                )
            if sample["context"] != reference["context"]:
                errors.append(
                    f"{sample['testcase_name']}: relation input context differs from "
                    f"{reference['testcase_name']}"
                )
            if sample["value"] != reference["value"]:
                errors.append(
                    f"{sample['testcase_name']}: raw output differs from "
                    f"{reference['testcase_name']}"
                )

        case_counts = Counter(sample["testcase_name"] for sample in samples)
        if self.require_intra_case:
            incomplete_cases = sorted(
                name for name, count in case_counts.items() if count < 2
            )
            if incomplete_cases:
                errors.append(
                    "same-case relation has fewer than two slices: "
                    + ", ".join(incomplete_cases)
                )
        if self.require_cross_case and len(case_counts) < 2:
            errors.append("relation appears in fewer than two testcases")

        return {
            "relation": list(relation),
            "status": "PASS" if not errors else "FAIL",
            "case_count": len(case_counts),
            "sample_count": len(samples),
            "samples": [
                {
                    "testcase_name": sample["testcase_name"],
                    "slice": sample["slice"],
                    "shape": sample["shape"],
                    "dtype": sample["dtype"],
                    "sha256": sample["digest"],
                }
                for sample in samples
            ],
            "errors": errors,
        }

    def write_excel(self, report):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        summary.append(("Field", "Value"))
        for key in (
            "status",
            "case_csv",
            "result_csv",
            "dump_dir",
            "output_index",
            "require_intra_case",
            "require_cross_case",
            "group_count",
            "sample_count",
        ):
            summary.append((key, str(report[key])))

        details = workbook.create_sheet("Relations")
        headers = (
            "Group",
            "Output",
            "Logical axes",
            "Seed",
            "Logical size",
            "Status",
            "Case count",
            "Sample count",
            "Testcase",
            "Slice",
            "Shape",
            "Dtype",
            "SHA-256",
            "Errors",
        )
        details.append(headers)
        for group_id, group in enumerate(report["groups"], start=1):
            first_row = details.max_row + 1
            output_index, batch_axis, seed, slice_length = group["relation"]
            errors = "\n".join(group["errors"])
            for sample in group["samples"]:
                details.append(
                    (
                        group_id,
                        output_index,
                        repr(batch_axis),
                        seed,
                        repr(slice_length),
                        group["status"],
                        group["case_count"],
                        group["sample_count"],
                        sample["testcase_name"],
                        repr(sample["slice"]),
                        repr(sample["shape"]),
                        sample["dtype"],
                        sample["sha256"],
                        errors,
                    )
                )
            last_row = details.max_row
            if last_row > first_row:
                for column in (*range(1, 9), 14):
                    details.merge_cells(
                        start_row=first_row,
                        start_column=column,
                        end_row=last_row,
                        end_column=column,
                    )
            for column in (*range(1, 9), 14):
                details.cell(first_row, column).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            fill = "C6EFCE" if group["status"] == "PASS" else "FFC7CE"
            details.cell(first_row, 6).fill = PatternFill("solid", fgColor=fill)

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for worksheet in (summary, details):
            worksheet.freeze_panes = "A2"
            worksheet.sheet_view.showGridLines = False
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
        details.freeze_panes = "I2"
        details.auto_filter.ref = details.dimensions
        for column, width in {
            "A": 9,
            "B": 9,
            "C": 8,
            "D": 12,
            "E": 14,
            "F": 10,
            "G": 12,
            "H": 14,
            "I": 54,
            "J": 20,
            "K": 24,
            "L": 16,
            "M": 68,
            "N": 58,
        }.items():
            details.column_dimensions[column].width = width
        summary.column_dimensions["A"].width = 28
        summary.column_dimensions["B"].width = 100

        self.excel_path.parent.mkdir(parents=True, exist_ok=True)
        with tempfile.NamedTemporaryFile(
            prefix=f".{self.excel_path.stem}.",
            suffix=".xlsx",
            dir=self.excel_path.parent,
            delete=False,
        ) as stream:
            temporary = Path(stream.name)
        try:
            workbook.save(temporary)
            os.replace(temporary, self.excel_path)
        finally:
            temporary.unlink(missing_ok=True)

    def run(self):
        if self.output_index != self.PRIMARY_OUTPUT_INDEX:
            raise ValueError("QSMLA batch consistency supports output index 0 only")
        case_rows = self.read_rows(self.case_csv)
        result_rows = self.read_rows(self.result_csv)
        disabled_cases = [
            row.get("testcase_name", "<unknown>")
            for row in case_rows
            if not self.case_is_enabled(row)
        ]
        enabled_case_rows = [row for row in case_rows if self.case_is_enabled(row)]
        samples = self.build_samples(enabled_case_rows, result_rows)
        grouped_samples = defaultdict(list)
        for sample in samples:
            grouped_samples[sample["relation"]].append(sample)
        groups = [
            self.compare_group(relation, values)
            for relation, values in sorted(grouped_samples.items())
        ]
        passed = all(group["status"] == "PASS" for group in groups)
        report = {
            "status": "PASS" if passed else "FAIL",
            "case_csv": str(self.case_csv.resolve()),
            "result_csv": str(self.result_csv.resolve()),
            "dump_dir": str(self.dump_dir.resolve()),
            "output_index": self.output_index,
            "require_intra_case": self.require_intra_case,
            "require_cross_case": self.require_cross_case,
            "group_count": len(groups),
            "sample_count": len(samples),
            "disabled_case_count": len(disabled_cases),
            "disabled_cases": disabled_cases,
            "groups": groups,
        }
        self.report_path.parent.mkdir(parents=True, exist_ok=True)
        self.report_path.write_text(
            json.dumps(report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
        )
        if self.excel_path is not None:
            self.write_excel(report)
        return passed, report


def build_parser():
    parser = argparse.ArgumentParser(
        description="Compare raw E2E output bins for batch-consistency relations."
    )
    parser.add_argument("--csv", required=True, help="E2E case CSV")
    parser.add_argument("--result", required=True, help="TTK result CSV")
    parser.add_argument(
        "--dump-dir", required=True, help="Directory set through NPU_DUMP_PATH"
    )
    parser.add_argument("--report", required=True, help="JSON report path")
    parser.add_argument("--excel", help="Optional grouped Excel report path")
    parser.add_argument(
        "--output-index", type=int, default=0, help="Output index dumped by E2E"
    )
    parser.add_argument(
        "--require-intra-case",
        action="store_true",
        help="Require every testcase to provide at least two slices for each relation",
    )
    parser.add_argument(
        "--require-cross-case",
        action="store_true",
        help="Require every relation to occur in at least two testcases",
    )
    return parser


def main():
    args = build_parser().parse_args()
    comparator = BatchOutputComparator(
        args.csv,
        args.result,
        args.dump_dir,
        args.report,
        args.output_index,
        args.require_intra_case,
        args.require_cross_case,
        args.excel,
    )
    try:
        passed, report = comparator.run()
    except (OSError, ValueError, csv.Error) as error:
        print(f"batch consistency comparison failed: {error}")
        return 1
    print(
        f"batch consistency {report['status']}: groups={report['group_count']}, "
        f"samples={report['sample_count']}, "
        f"skipped_disabled={report['disabled_case_count']}, report={args.report}"
    )
    return 0 if passed else 1


if __name__ == "__main__":
    raise SystemExit(main())
