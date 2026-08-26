#!/usr/bin/python3
# -*- coding: utf-8 -*-
# -----------------------------------------------------------------------------------------------------------
# Copyright (c) 2026 Huawei Technologies Co., Ltd.
# This program is free software, you can redistribute it and/or modify it under the terms and conditions of
# CANN Open Software License Agreement Version 2.0 (the "License").
# Please refer to the License for details. You may not use this file except in compliance with the License.
# THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND, EITHER EXPRESS OR IMPLIED,
# INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT, MERCHANTABILITY, OR FITNESS FOR A PARTICULAR PURPOSE.
# See LICENSE in the root of the software repository for the full text of the License.
# -----------------------------------------------------------------------------------------------------------
"""极简 xlsx 读取器：zipfile + xml.dom.minidom，无 openPyXL/pandas 依赖。

供 excel_to_csv.py 使用。默认读 xl/worksheets/sheet1.xml +
xl/sharedStrings.xml，返回 list[dict[col_idx]→text]，空单元格 → None。

支持指定 sheet:
  - load_sheet1(path)          → 读 sheet1 (向后兼容)
  - load_sheet(path, sheet)    → sheet 可为序号 (1→sheet1.xml) 或 sheet 名
    (从 xl/workbook.xml 的 <sheet name=... r:id=...> 映射到 xl/worksheets/sheetN.xml)
"""

import re
import zipfile
import xml.dom.minidom as minidom

_CELL_RE = re.compile(r"^([A-Z]+)(\d+)$")


def col_to_idx(col: str) -> int:
    """列字母 → 0-based 列索引（A=0 ... Z=25, AA=26, AB=27 ...）"""
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch.upper()) - ord("A") + 1)
    return n - 1


def header_to_name_idx(header_row: dict) -> dict:
    """表头行 (dict[col_idx→列名]) 反转为 dict[列名→col_idx]。

    供 excel_to_csv.py 按列名动态定位列, 避免不同 sheet 列顺序偏移导致错位。
    重复列名取最后一个出现的 col_idx (罕见, 保留警告语义由调用方处理)。
    空列名 (None/空串) 跳过。
    """
    return {name: idx for idx, name in sorted(header_row.items()) if name}


def _parse_ref(ref: str):
    """把 Excel 单元格引用（如 "AB12"）拆成 (col_idx, row_idx)。"""
    m = _CELL_RE.match(ref)
    if not m:
        raise ValueError(f"bad cell ref: {ref!r}")
    return col_to_idx(m.group(1)), int(m.group(2)) - 1


def _cell_text(cell, shared_strings):
    """取一个 <c> 元素的文本值。返回 None 表示空单元格。

    - t="s" + <v>i</v>      → shared_strings[i]
    - t="inlineStr" + <is><t> → inline string
    - t="b" + <v>0/1</v>    → bool 字符串 "True"/"False"
    - <v>num</v>            → 数值字符串（保留 Excel 原样，避免精度丢失）
    - <f>...</f> + <v>      → 公式缓存值，取 <v>
    - 空                    → None
    """
    t = cell.getAttribute("t")
    v_els = cell.getElementsByTagName("v")
    if t == "s" and v_els:
        idx = int(v_els[0].firstChild.data)
        return shared_strings[idx]
    if t == "b" and v_els:
        return "True" if v_els[0].firstChild.data == "1" else "False"
    if t == "inlineStr":
        is_els = cell.getElementsByTagName("is")
        if not is_els:
            return None
        ts = is_els[0].getElementsByTagName("t")
        return "".join(tt.firstChild.data if tt.firstChild else "" for tt in ts)
    if v_els:
        # 数值或公式缓存值
        return v_els[0].firstChild.data
    return None


def _resolve_sheet_xml(z: zipfile.ZipFile, sheet) -> str:
    """把 sheet 标识 (int 序号或 str 名) 解析为 xl/worksheets/sheetN.xml 路径。

    - sheet 为 int: 直接作为 sheet 序号 (1 → sheet1.xml, 2 → sheet2.xml ...)
    - sheet 为 str: 从 xl/workbook.xml 的 <sheet name=... r:id="rIdN"/> 查询,
      再从 xl/_rels/workbook.xml.rels 的 <Relationship Id="rIdN" Target="worksheets/sheetN.xml"/>
      解析实际文件路径
    """
    if isinstance(sheet, int):
        return f"xl/worksheets/sheet{sheet}.xml"

    # 按名查询: workbook.xml → r:id → rels → worksheets/sheetN.xml
    wb_xml = z.read("xl/workbook.xml").decode("utf-8")
    wb_doc = minidom.parseString(wb_xml)
    rid = None
    for s in wb_doc.getElementsByTagName("sheet"):
        name = s.getAttribute("name")
        if name == sheet:
            rid = s.getAttribute("r:id") or s.getAttribute(
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
            )
            break
    if rid is None:
        # 列出所有可用 sheet 名便于排错
        names = [s.getAttribute("name") for s in wb_doc.getElementsByTagName("sheet")]
        raise ValueError(
            f"sheet name {sheet!r} not found in workbook. available: {names}"
        )

    rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    rels_doc = minidom.parseString(rels_xml)
    for rel in rels_doc.getElementsByTagName("Relationship"):
        if rel.getAttribute("Id") == rid:
            target = rel.getAttribute("Target")
            # Target 形如 "worksheets/sheet2.xml" (相对 xl/) 或绝对路径
            if target.startswith("/"):
                return target.lstrip("/")
            return "xl/" + target
    raise ValueError(f"r:id {rid!r} (sheet {sheet!r}) not found in workbook.xml.rels")


def _parse_sheet_xml(sheet_xml: str, shared_strings: list) -> list:
    """解析单个 sheet xml 为 list[dict[col_idx]→text]"""
    sh = minidom.parseString(sheet_xml)
    rows = sh.getElementsByTagName("row")
    rows = sorted(rows, key=lambda r: int(r.getAttribute("r")))
    parsed_rows = []
    for row in rows:
        cells = {}
        for c in row.getElementsByTagName("c"):
            ref = c.getAttribute("r")
            col_idx, _row_idx = _parse_ref(ref)
            cells[col_idx] = _cell_text(c, shared_strings)
        parsed_rows.append(cells)
    return parsed_rows


def load_sheet(xlsx_path: str, sheet=1):
    """读取指定 sheet。sheet 可为序号 (int, 1-based) 或 sheet 名 (str)。

    优先用 zipfile + minidom 解析 (无第三方依赖); 若 xlsx 不含
    xl/worksheets/sheetN.xml (罕见压缩结构) 回退到 openpyxl。
    """
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            ss_xml = z.read("xl/sharedStrings.xml").decode("utf-8")
            sheet_path = _resolve_sheet_xml(z, sheet)
            sheet_xml = z.read(sheet_path).decode("utf-8")
    except (KeyError, FileNotFoundError):
        return _load_sheet_openpyxl(xlsx_path, sheet)

    ss_doc = minidom.parseString(ss_xml)
    shared_strings = []
    for si in ss_doc.getElementsByTagName("si"):
        ts = si.getElementsByTagName("t")
        shared_strings.append(
            "".join(t.firstChild.data if t.firstChild else "" for t in ts)
        )
    return _parse_sheet_xml(sheet_xml, shared_strings)


def load_sheet1(xlsx_path: str):
    """读 sheet1 (向后兼容别名, 等价于 load_sheet(xlsx_path, 1))"""
    return load_sheet(xlsx_path, 1)


def _load_sheet_openpyxl(xlsx_path: str, sheet=1):
    """openpyxl 回退路径, 支持 sheet 序号或名"""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if isinstance(sheet, int):
        # 1-based → 第 N 个 sheet
        ws = wb.worksheets[sheet - 1] if 1 <= sheet <= len(wb.worksheets) else wb.active
    else:
        ws = wb[sheet]
    parsed_rows = []
    for row in ws.iter_rows():
        cells = {}
        for cell in row:
            v = cell.value
            if v is None:
                continue
            cells[cell.col_idx - 1] = str(v) if not isinstance(v, str) else v
        parsed_rows.append(cells)
    return parsed_rows


def _load_sheet1_openpyxl(xlsx_path: str):
    return _load_sheet_openpyxl(xlsx_path, 1)


def list_sheet_names(xlsx_path: str) -> list:
    """返回 redline.xlsx 中所有 sheet 名 (供用户查看可选 sheet)"""
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            wb_xml = z.read("xl/workbook.xml").decode("utf-8")
    except (KeyError, FileNotFoundError):
        import openpyxl

        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        return wb.sheetnames
    wb_doc = minidom.parseString(wb_xml)
    return [s.getAttribute("name") for s in wb_doc.getElementsByTagName("sheet")]
