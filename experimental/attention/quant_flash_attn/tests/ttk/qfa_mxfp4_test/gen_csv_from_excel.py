#!/usr/bin/python3
# -*- coding: utf-8 -*-
# -----------------------------------------------------------------------------------------------------------
# Copyright (c) 2026 Huawei Technologies Co., Ltd.
# This program is free software, you can redistribute it and/or modify it under the terms and conditions of
# CANN Open Software License Agreement Version 2.0 (the "License").
# Please refer to the License for details. You may not use this file except in compliance with the License.
# THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND, EITHER EXPRESS OR IMPLIED,
# INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT, MERCHANTABILITY, OR FITNESS FOR A PARTICULAR PURPOSE.
# See LICENSE in the root of the software repository for the full text of the License.
# -----------------------------------------------------------------------------------------------------------
"""
python3 gen_csv_from_excel.py [--excel B008QFA_红线用例.xlsx]
                              [--sheet mxfp4]
                              [--output qfa_mxfp4.csv]

"""

import argparse
import csv
import os
import sys

import openpyxl

_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)

API_NAME = "qfa_mxfp4_wrapper.npu_qfa_mxfp4"

# 字段 key -> Excel 表头名 (列位置无关, 启动时按表头名动态查列号)
# 注意: Excel C-I 列 (B/Q_N/KV_N/Q_S/KV_S/D/G) 是给开发人员看的汇总信息, 不读取;
#       实际从 batch_size/num_heads_q/num_heads_kv/head_dim (BU-CC 列) 取值.
HEADER = {
    "name": "testcase_name",
    # 基本信息 (从 J 列起, 跳过 C-I 汇总列)
    "out_dtype": "attn_out_dtype",
    # q/k/v 数据
    "q_shape": "q_shape",
    "q_dtype": "q_dtype",
    "q_datarange": "q_datarange",
    "k_shape": "k_shape",
    "k_dtype": "k_dtype",
    "k_datarange": "k_datarange",
    "v_shape": "v_shape",
    "v_dtype": "v_dtype",
    "v_datarange": "v_datarange",
    # q/k/v descale
    "q_descale_shape": "q_descale_shape",
    "q_descale_dtype": "q_descale_dtype",
    "k_descale_shape": "k_descale_shape",
    "k_descale_dtype": "k_descale_dtype",
    "v_descale_shape": "v_descale_shape",
    "v_descale_dtype": "v_descale_dtype",
    # 可选 tensor 入参 (shape 为空 -> golden 传 None)
    "block_table_shape": "block_table_shape",
    "block_table_dtype": "block_table_dtype",
    "p_scale_value": "p_scale_value",
    "p_scale_shape": "p_scale_shape",
    "p_scale_dtype": "p_scale_dtype",
    "p_scale_datarange": "p_scale_datarange",
    "learnable_sink_shape": "learnable_sink_shape",
    "learnable_sink_dtype": "learnable_sink_dtype",
    "learnable_sink_datarange": "learnable_sink_datarange",
    "attn_mask_shape": "attn_mask_shape",
    "attn_mask_dtype": "attn_mask_dtype",
    "attn_mask_datarange": "attn_mask_datarange",
    # seq 信息
    "cu_seqlens_q_value": "cu_seqlens_q_value",
    "cu_seqlens_q_dtype": "cu_seqlens_q_dtype",
    "cu_seqlens_kv_value": "cu_seqlens_kv_value",
    "cu_seqlens_kv_dtype": "cu_seqlens_kv_dtype",
    "seqused_q_value": "seqused_q_value",
    "seqused_q_dtype": "seqused_q_dtype",
    "seqused_kv_value": "seqused_kv_value",
    "seqused_kv_dtype": "seqused_kv_dtype",
    # softmax_lse 输出 dtype
    "softmax_lse_dtype": "softmax_lse_dtype",
    # 算子配置
    "quant_mode": "quant_mode",
    "softmax_scale": "softmax_scale",
    "mask_mode": "mask_mode",
    "win_left": "win_left",
    "win_right": "win_right",
    "max_seqlen_q": "max_seqlen_q",
    "max_seqlen_kv": "max_seqlen_kv",
    # layout
    "layout_q": "layout_q",
    "layout_q_descale": "layout_q_descale",
    "layout_kv": "layout_kv",
    "layout_out": "layout_out",
    # 开关
    "return_softmax_lse": "return_softmax_lse",
    # 维度 (从 BU-CC 列取, 不从 C-I 汇总列取)
    "batch_size": "batch_size",
    "num_heads_q": "num_heads_q",
    "num_heads_kv": "num_heads_kv",
    "head_dim": "head_dim",
}

# 可选精度列 (缺失时用默认值)
OPTIONAL_HEADER = ("precision_tolerances", "absolute_precision")


def _parse_int_list(s):
    """'1024,1022,1023' -> [1024, 1022, 1023]. None/空 -> []."""
    if s is None:
        return []
    s = str(s).strip()
    if not s:
        return []
    return [int(x.strip()) for x in s.split(",") if x.strip() != ""]


def _parse_shape(s):
    """'8,12,1024,128' -> (8,12,1024,128)."""
    return tuple(_parse_int_list(s))


def _norm_dtype(v):
    """Excel dtype 字符串 -> wrapper 期望的 dtype 名."""
    if v is None:
        return None
    s = str(v).strip().upper()
    if s in ("FP4_E2M1", "FP4E2M1"):
        return "fp4_e2m1"
    if s in ("BF16",):
        return "bfloat16"
    if s in ("FLOAT8_E8M0",):
        return "uint8"  # E8M0 scale 用 uint8 存
    return str(v).strip().lower()


def _compute_v_descale_shape(B, N_kv, V_D, act_seq_lens_kv, max_seqlen_kv=-1):
    """按 golden 量化逻辑重算 v_descale shape (BNSD).

    s2 = max(act_seq_lens_kv) if 有 seqused, else max_seqlen_kv; n_blocks = ceil(s2/32); 奇数 pad +1
    最终 (B, N_kv, n_blocks // 2, V_D, 2).
    用于校验 Excel v_descale 列 (Excel 偶用 v_shape 的 S 而非 max(seqused_kv) 致错).
    """
    s2 = (
        max(act_seq_lens_kv)
        if act_seq_lens_kv
        else (max_seqlen_kv if max_seqlen_kv >= 0 else 0)
    )
    n_blocks = (s2 + 31) // 32
    if n_blocks % 2 != 0:
        n_blocks += 1
    return (B, N_kv, n_blocks // 2, V_D, 2)


def _half_last_dim(shape):
    """MXFP4 packed: 最后一维 //2."""
    return tuple(list(shape[:-1]) + [shape[-1] // 2])


# 精度阈值默认值 (bfloat16 输出): 两个输出各一个 (rtol, ptol) 对 + 各一个 atol
DEFAULT_PRECISION_TOLERANCES = "((0.0078125, 0.005), (0.0078125, 0.005))"
DEFAULT_ABSOLUTE_PRECISION = "(0.0001, 0.0001)"


def _find_col_by_header(ws, header_name, start_col=1):
    """在表头行 (row 1) 按列名查找列号 (1-indexed). 找不到返回 None."""
    for c in range(start_col, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip() == header_name:
            return c
    return None


def _build_col_map(ws):
    """扫描表头行, 构建 字段key -> 列号 映射.

    必需列 (HEADER) 缺失则抛错并列出缺哪些; 可选列 (OPTIONAL_HEADER) 缺失则不加入 map.
    """
    col_map = {}
    missing = []
    for key, hdr_name in HEADER.items():
        col = _find_col_by_header(ws, hdr_name)
        if col is None:
            missing.append(f"{key}='{hdr_name}'")
        else:
            col_map[key] = col
    if missing:
        raise ValueError(
            f"Excel sheet 缺少必需列 (共 {len(missing)} 个): {', '.join(missing)}\n"
            f"请保证表头名与脚本 HEADER 定义一致 (列位置可任意)."
        )
    for hdr_name in OPTIONAL_HEADER:
        col = _find_col_by_header(ws, hdr_name)
        if col is not None:
            col_map[hdr_name] = col
    return col_map


def excel_row_to_csv_row(ws, row_idx, col_map):
    """读 Excel 一行 -> CSV 行 (10 列).

    col_map: _build_col_map 返回的 字段key -> 列号 映射 (含可选精度列).
    """

    def g(key):
        return ws.cell(row=row_idx, column=col_map[key]).value

    def g_opt(key):
        col = col_map.get(key)
        return ws.cell(row=row_idx, column=col).value if col is not None else None

    name = str(g("name")).strip()
    B = int(g("batch_size"))
    N_q = int(g("num_heads_q"))
    N_kv = int(g("num_heads_kv"))
    D = int(g("head_dim"))
    G = N_q // N_kv if N_kv > 0 else 1  # GQA group size = Q heads // KV heads
    V_D = D  # Excel head_dim == D, 无独立 V_D 列

    # act_seq_lens_q/kv: 透明透传 Excel 的 seqused 列; 为空则传空
    sq = _parse_int_list(g("seqused_q_value"))
    skv = _parse_int_list(g("seqused_kv_value"))

    # cu_seqlens_q/kv: 透明透传 (TND 才有值, 否则空)
    cu_sq = _parse_int_list(g("cu_seqlens_q_value"))
    cu_skv = _parse_int_list(g("cu_seqlens_kv_value"))

    # max_seqlen_q/kv: 透明透传 Excel 的 max_seqlen 列; 缺失/空 -> -1 (算子侧 -1 表示不传)
    ms_q_raw = g("max_seqlen_q")
    ms_kv_raw = g("max_seqlen_kv")
    max_seqlen_q = (
        int(ms_q_raw) if ms_q_raw is not None and str(ms_q_raw).strip() != "" else -1
    )
    max_seqlen_kv = (
        int(ms_kv_raw) if ms_kv_raw is not None and str(ms_kv_raw).strip() != "" else -1
    )

    # pre/next_tokens: -1 -> 2147483647
    win_left = g("win_left")
    pre_tokens = (
        2147483647 if (win_left is None or int(win_left) == -1) else int(win_left)
    )
    win_right = g("win_right")
    next_tokens = (
        2147483647 if (win_right is None or int(win_right) == -1) else int(win_right)
    )

    # q/k/v: Excel 是 unpacked shape (末维 = D), MXFP4 packed -> 末维 //2
    q_shape = _half_last_dim(_parse_shape(g("q_shape")))
    k_shape = _half_last_dim(_parse_shape(g("k_shape")))
    v_shape = _half_last_dim(_parse_shape(g("v_shape")))

    # q/k_descale: 直接取 Excel shape
    q_descale_shape = _parse_shape(g("q_descale_shape"))
    k_descale_shape = _parse_shape(g("k_descale_shape"))

    # v_descale: golden 重算校验, 不一致则用 golden 并告警
    v_descale_excel = _parse_shape(g("v_descale_shape"))
    v_descale_golden = _compute_v_descale_shape(B, N_kv, V_D, skv, max_seqlen_kv)
    if v_descale_excel != v_descale_golden:
        print(
            f"[WARN] {name}: Excel v_descale={v_descale_excel} 与 golden={v_descale_golden} "
            f"不一致, 使用 golden (Excel 公式疑似用 v_shape 的 S 而非 max(seqused_kv))"
        )
    v_descale_shape = v_descale_golden

    block_table_shape = (
        0,
    )  # TTK 占位 (continue KV), 真实 block_table 从 attrs 读 shape 生成

    shapes = [
        q_shape,
        k_shape,
        v_shape,
        q_descale_shape,
        k_descale_shape,
        v_descale_shape,
        block_table_shape,
    ]
    tensor_view_shapes = "(" + ",".join(f"({s})" for s in shapes) + ")"

    dtypes = [
        "'uint8'",
        "'uint8'",
        "'uint8'",
        "'uint8'",
        "'uint8'",
        "'uint8'",
        "'int32'",
    ]
    tensor_dtypes = "(" + ",".join(dtypes) + ")"

    # layout: 从 Excel 读取 (layout_q / layout_q_descale / layout_kv / layout_out)
    layout_q = str(g("layout_q")).strip()
    layout_q_descale = str(g("layout_q_descale")).strip()
    layout_kv = str(g("layout_kv")).strip()
    layout_out = str(g("layout_out")).strip()

    # 可选 tensor 入参: 从 Excel 读 shape, 有 shape 则 golden 生成随机 tensor, 无 shape 传 None
    def _read_opt_tensor(shape_key, dtype_key=None, datarange_key=None):
        """读 Excel 可选 tensor 的 shape/dtype/datarange, 返回 dict (无值则 shape=())"""
        sh = _parse_shape(g(shape_key))
        dt = _norm_dtype(g(dtype_key)) if dtype_key else None
        dr = (
            str(g(datarange_key))
            if datarange_key and g(datarange_key) is not None
            else None
        )
        return {"shape": list(sh), "dtype": dt, "datarange": dr}

    block_table_info = _read_opt_tensor("block_table_shape", "block_table_dtype")
    p_scale_info = _read_opt_tensor(
        "p_scale_shape", "p_scale_dtype", "p_scale_datarange"
    )
    sinks_info = _read_opt_tensor(
        "learnable_sink_shape", "learnable_sink_dtype", "learnable_sink_datarange"
    )
    attn_mask_info = _read_opt_tensor(
        "attn_mask_shape", "attn_mask_dtype", "attn_mask_datarange"
    )

    # p_scale_value: 表格传了就用表格值 (标量), 否则 None
    p_scale_value = g("p_scale_value")
    if p_scale_value is not None and str(p_scale_value).strip() != "":
        try:
            p_scale_value = float(p_scale_value)
        except (ValueError, TypeError):
            p_scale_value = None
    else:
        p_scale_value = None

    # enable_mask: attn_mask_shape 有值 -> True, 否则 False
    enable_mask = len(attn_mask_info["shape"]) > 0

    attrs = {
        "B": B,
        "N_q": N_q,
        "N_kv": N_kv,
        "G": G,
        "D": D,
        "V_D": V_D,
        "Rope_D": 0,
        "act_seq_lens_q": sq,
        "act_seq_lens_kv": skv,
        "cu_seqlens_q": cu_sq,
        "cu_seqlens_kv": cu_skv,
        "max_seqlen_q": max_seqlen_q,
        "max_seqlen_kv": max_seqlen_kv,
        "input_layout": layout_q,
        "layout_q_descale": layout_q_descale,
        "layout_kv": layout_kv,
        "layout_out": layout_out,
        "kv_storage_mode": "continue",
        "block_size": 0,
        "q_dtype": _norm_dtype(g("q_dtype")) or "fp4_e2m1",
        "kv_dtype": _norm_dtype(g("k_dtype")) or "fp4_e2m1",
        "out_dtype": _norm_dtype(g("out_dtype")) or "bfloat16",
        "q_quant_mode": int(g("quant_mode")) if g("quant_mode") is not None else 3,
        "mask_mode": int(g("mask_mode")) if g("mask_mode") is not None else 0,
        "pre_tokens": pre_tokens,
        "next_tokens": next_tokens,
        "enable_mask": enable_mask,
        "enable_lse": bool(g("return_softmax_lse"))
        if g("return_softmax_lse") is not None
        else False,
        "inner_precise": 0,
        "device_id": 0,
        "graph_path": 0,
        "softmax_scale": g("softmax_scale"),
        # 可选 tensor 入参 (shape 为空 -> golden 传 None)
        "block_table_shape": block_table_info["shape"],
        "block_table_dtype": block_table_info["dtype"],
        "p_scale_value": p_scale_value,
        "p_scale_shape": p_scale_info["shape"],
        "p_scale_dtype": p_scale_info["dtype"],
        "p_scale_datarange": p_scale_info["datarange"],
        "sinks_shape": sinks_info["shape"],
        "sinks_dtype": sinks_info["dtype"],
        "sinks_datarange": sinks_info["datarange"],
        "attn_mask_shape": attn_mask_info["shape"],
        "attn_mask_dtype": attn_mask_info["dtype"],
        "attn_mask_datarange": attn_mask_info["datarange"],
        # dtype 透传 (表格不传 -> None, golden 侧用默认值)
        "q_descale_dtype": _norm_dtype(g("q_descale_dtype")),
        "k_descale_dtype": _norm_dtype(g("k_descale_dtype")),
        "v_descale_dtype": _norm_dtype(g("v_descale_dtype")),
        "seqused_q_dtype": _norm_dtype(g("seqused_q_dtype")),
        "seqused_kv_dtype": _norm_dtype(g("seqused_kv_dtype")),
        "cu_seqlens_q_dtype": _norm_dtype(g("cu_seqlens_q_dtype")),
        "cu_seqlens_kv_dtype": _norm_dtype(g("cu_seqlens_kv_dtype")),
        "softmax_lse_dtype": _norm_dtype(g("softmax_lse_dtype")),
        "data_range_q": str(g("q_datarange")) if g("q_datarange") is not None else 1.0,
        "data_range_k": str(g("k_datarange")) if g("k_datarange") is not None else 1.0,
        "data_range_v": str(g("v_datarange")) if g("v_datarange") is not None else 1.0,
    }
    attributes = repr(attrs)

    # 精度阈值: Excel 有对应列则原样透传字符串, 否则用默认值
    pt_raw = g_opt("precision_tolerances")
    ap_raw = g_opt("absolute_precision")
    precision_tolerances = (
        str(pt_raw).strip() if pt_raw is not None else DEFAULT_PRECISION_TOLERANCES
    )
    absolute_precision = (
        str(ap_raw).strip() if ap_raw is not None else DEFAULT_ABSOLUTE_PRECISION
    )

    return [
        name,
        API_NAME,
        tensor_view_shapes,
        tensor_dtypes,
        "",
        attributes,
        "",
        "",
        precision_tolerances,
        absolute_precision,
    ]


def main():
    parser = argparse.ArgumentParser(
        description="Excel mxfp4 sheet -> TTK e2e CSV (批量转换交付件)"
    )
    parser.add_argument("--excel", default="B008QFA_红线用例.xlsx", help="Excel 路径")
    parser.add_argument("--sheet", default="mxfp4", help="sheet 名 (默认 mxfp4)")
    parser.add_argument("--output", default="qfa_mxfp4.csv", help="输出 CSV 路径")
    args = parser.parse_args()

    excel_path = (
        args.excel if os.path.isabs(args.excel) else os.path.join(_HERE, args.excel)
    )
    out_path = (
        args.output if os.path.isabs(args.output) else os.path.join(_HERE, args.output)
    )

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"sheet '{args.sheet}' 不存在, 可用: {wb.sheetnames}")
    ws = wb[args.sheet]

    # 按表头名动态构建 字段key -> 列号 映射
    col_map = _build_col_map(ws)
    prec_found = {h: col_map[h] for h in OPTIONAL_HEADER if h in col_map}
    if prec_found:
        print(f"[INFO] 检测到 Excel 精度列: {prec_found} (从 Excel 读取)")
    else:
        print(
            f"[INFO] Excel 无精度列, 使用默认值: "
            f"precision_tolerances={DEFAULT_PRECISION_TOLERANCES}, "
            f"absolute_precision={DEFAULT_ABSOLUTE_PRECISION}"
        )

    header = [
        "testcase_name",
        "api_name",
        "tensor_view_shapes",
        "tensor_dtypes",
        "tensor_formats",
        "attributes",
        "output_tensor_indexes",
        "golden_api",
        "precision_tolerances",
        "absolute_precision",
    ]
    rows = []
    name_col = col_map["name"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=name_col).value is None:
            continue
        rows.append(excel_row_to_csv_row(ws, r, col_map))
        print(f"[Excel] {rows[-1][0]}")

    with open(out_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for row in rows:
            writer.writerow(row)

    print(f"\n生成 {len(rows)} 条用例 -> {out_path}")


if __name__ == "__main__":
    main()
